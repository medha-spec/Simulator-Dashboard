"""
V1.2 data pipeline for the Snitch Control Room dashboard.

THREE ways to run this:

  A) Fully automatic, authenticated, IMPORTRANGE-safe (recommended -- set up once):
       python build_data.py
     Requires a Google Service Account (one-time ~10 min setup, see README).
     Save its JSON key as ./service_account.json next to this script, and share
     the Google Sheet with the service account's email (Viewer access).
     This goes through Google's live Sheets API, which correctly resolves
     IMPORTRANGE (and any other cross-sheet formula) because it's a real
     authenticated read, unlike a plain anonymous URL fetch.

  B) Automatic but UNSAFE for IMPORTRANGE-driven tabs:
       python build_data.py --export-url
     Downloads via the plain https://.../export?format=xlsx URL. Fine for
     Sales Data and Pipeline (plain formulas), but Inv Data 2 (IMPORTRANGE)
     will come back BLANK -- Google's export endpoint can't resolve
     IMPORTRANGE without an authenticated session. Kept only as a fallback.

  C) Offline / manual:
     python build_data.py --local
     Reads ./Automation_Data.xlsx if you've placed one there yourself.

Either way it writes data/data.json, which is the only thing index.html reads.

Sales now comes from Snowflake (SALES_FOR_AUTO_2, daily grain), not the "Sales
Data" Google Sheet tab. Requires these env vars (or a local
snowflake_credentials.json -- see snowflake_credentials.example.json):
  SNOWFLAKE_ACCOUNT, SNOWFLAKE_USER, SNOWFLAKE_PASSWORD,
  SNOWFLAKE_WAREHOUSE, SNOWFLAKE_DATABASE, SNOWFLAKE_SCHEMA
Inventory and Pipeline still come from Google Sheets as before.

Setup for mode A (one-time):
  pip install google-api-python-client google-auth --break-system-packages
  1. console.cloud.google.com -> new/existing project -> enable "Google Sheets API"
  2. IAM & Admin -> Service Accounts -> Create Service Account (any name, no roles needed)
  3. That service account -> Keys -> Add Key -> Create new key -> JSON -> downloads a file
  4. Save that file as service_account.json in this same folder
  5. Open the file, copy the "client_email" value
  6. Share the actual Google Sheet with that email, Viewer access
  Done -- python build_data.py now works with zero manual downloads, forever.
"""
import json, datetime, sys, os, csv
import openpyxl

# Rolling window for Snowflake-sourced data (Sales, Returns actual, Sales-by-SKU).
# Computed fresh every run from *today*, not a fixed date -- this is what keeps
# sales.json/returns_actual.json/sales_by_sku.json from growing bigger every
# single day forever. 400 days covers: current month, full prior-year MoM
# comparison, and any reasonable Custom Range pick in the Analytics tab, while
# old days automatically age out of the window as time moves forward.
ROLLING_WINDOW_DAYS = 400
ROLLING_CUTOFF_DATE = (datetime.date.today() - datetime.timedelta(days=ROLLING_WINDOW_DAYS)).isoformat()

SHEET_ID = "1PqtpL9w2Tneon_-6zz7BGiW5YUz4fhDCrgn687r_CZw"
HERE = os.path.dirname(os.path.abspath(__file__))
LOCAL_XLSX = os.path.join(HERE, "Automation_Data.xlsx")
SERVICE_ACCOUNT_FILE = os.path.join(HERE, "service_account.json")
TARGETS_DIR = os.path.join(HERE, "data", "targets")
OUT = os.path.join(HERE, "data", "data.json")
TAB_NAMES = ["Inv Data 2", "Pipeline"]  # Sales Data removed -- now sourced from Snowflake

# Exact filenames expected in data/targets/ -- one per category family, each with
# a "Sheet1" tab at store/branch grain with monthly qty columns (AUG_2026.. JUN_2027).
TARGET_FILES = [
    "Jeans Planned Qty only.xlsx",
    "Shirts Planned Qty Only.xlsx",
    "Trousers Planned Qty only.xlsx",
    "tshirts Planned Qty Only.xlsx",
]
MONTH_COL_MAP = {
    "AUG_2026": "2026-08", "SEP_2026": "2026-09", "OCT_2026": "2026-10",
    "NOV_2026": "2026-11", "DEC_2026": "2026-12", "JAN_2027": "2027-01",
    "FEB_2027": "2027-02", "MAR_2027": "2027-03", "APR_2027": "2027-04",
    "MAY_2027": "2027-05", "JUN_2027": "2027-06",
}
# normalize target's channel labels to match Sales' channel naming where they're
# clearly the same thing; Warehouse kept as its own distinct channel since Sales
# doesn't have an equivalent yet. MP-SOR merged into Marketplace per instruction.
TARGET_CHANNEL_MAP = {"Online - Shopify": "Shopify"}
# NOTE: MP-SOR deliberately kept distinct from Marketplace here (not merged),
# per instruction -- the channel-mix split treats them as separate percentages.
# The live dashboard's "Marketplace" filter still sums MP + MP-SOR together at
# query time (same as it already does for Sales), so nothing visible changes.

NEW_TOTAL_FILE = os.path.join(HERE, "data", "targets_new_total.xlsx")


# ---------------------------------------------------------------------------
# Mode A: Google Sheets API via service account (authenticated, IMPORTRANGE-safe)
# ---------------------------------------------------------------------------
GSHEET_EPOCH = datetime.datetime(1899, 12, 30)

def _gsheet_serial_to_date(value):
    """Sheets API returns dates as serial-day floats (same epoch as Excel).
    Converts back to a python datetime so downstream code (month_key etc.)
    doesn't need to know the difference between this and an openpyxl cell."""
    if isinstance(value, (int, float)):
        return GSHEET_EPOCH + datetime.timedelta(days=value)
    return value


class SimpleSheet:
    """Minimal stand-in for an openpyxl worksheet, backed by raw API rows,
    so the same parsing code in build() works for either data source."""
    def __init__(self, rows, width, date_cols=()):
        # pad every row out to `width` columns (Sheets API omits trailing blanks)
        self.rows = []
        for row in rows:
            padded = list(row) + [None] * (width - len(row))
            for c in date_cols:
                if isinstance(padded[c], str) and padded[c] == "":
                    padded[c] = None
                elif isinstance(padded[c], (int, float)):
                    padded[c] = _gsheet_serial_to_date(padded[c])
            self.rows.append(tuple(padded))

    def iter_rows(self, min_row=1, max_row=None, values_only=True):
        start = min_row - 1
        end = max_row if max_row else len(self.rows)
        for r in self.rows[start:end]:
            yield r

    @property
    def max_row(self):
        return len(self.rows)


class SimpleWorkbook:
    def __init__(self, sheets_dict):
        self._sheets = sheets_dict
        self.sheetnames = list(sheets_dict.keys())

    def __getitem__(self, name):
        return self._sheets[name]


def fetch_via_sheets_api(sheet_id):
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        print(f"ERROR: {SERVICE_ACCOUNT_FILE} not found.")
        print("See the module docstring (top of this file) for one-time setup steps.")
        sys.exit(1)

    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE,
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"]
    )
    service = build("sheets", "v4", credentials=creds)
    values_api = service.spreadsheets().values()

    # date columns (0-indexed) per tab, so serials get converted correctly
    date_cols = {"Inv Data 2": (0,), "Pipeline": (0, 1, 20, 21)}
    widths = {"Inv Data 2": 17, "Pipeline": 28}

    sheets = {}
    for tab in TAB_NAMES:
        print(f"Fetching '{tab}' via Sheets API...")
        result = values_api.get(
            spreadsheetId=sheet_id, range=tab,
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        rows = result.get("values", [])
        sheets[tab] = SimpleSheet(rows, widths[tab], date_cols.get(tab, ()))
        print(f"  -> {len(rows)} rows (incl. header)")

    return SimpleWorkbook(sheets)


def download_from_gsheet(sheet_id, dest):
    import requests
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    print("Downloading workbook from Google Sheets (unauthenticated export)...")
    print("WARNING: this will NOT resolve IMPORTRANGE-driven tabs (e.g. Inv Data 2).")
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    with open(dest, "wb") as f:
        f.write(r.content)
    print(f"Saved -> {dest}")


# ---------------------------------------------------------------------------
# Snowflake: daily-grain Sales (SALES_FOR_AUTO_2), replaces the "Sales Data"
# Google Sheet tab. Credentials come from env vars so the same code works
# locally (via snowflake_credentials.json, loaded into env vars below) and
# in GitHub Actions (via repo secrets injected as env vars in the workflow).
# ---------------------------------------------------------------------------
SNOWFLAKE_CREDS_FILE = os.path.join(HERE, "snowflake_credentials.json")

SNOWFLAKE_QUERY = f"""
    SELECT DATE, CHANNEL, L1_CATEGORY, CATEGORY, META1, META2, META3,
           SUM(GROSS_SALES_VALUE) AS GROSS_SALES_VALUE, SUM(MRP_VALUE) AS MRP_VALUE,
           SUM(QTY) AS QTY, SUM(COGS_SOLD) AS COGS_SOLD
    FROM SNITCH_DB.MAPLEMONK.SALES_FOR_AUTO_3
    WHERE DATE >= '{ROLLING_CUTOFF_DATE}'
    GROUP BY DATE, CHANNEL, L1_CATEGORY, CATEGORY, META1, META2, META3
"""

# Separate, much smaller SKU-level dataset for the Pareto/Quartile tabs only --
# aggregated by MONTH (not by day), since those tabs only need period totals
# per sku_group, not daily granularity. Keeping this out of the main SNOWFLAKE_QUERY
# above is what keeps data.json from ballooning to 1GB+ (SALES_FOR_AUTO_3 is raw
# SKU-day grain; exploding that into the main sales array multiplies file size
# by roughly the average SKU count per leaf combo).
SALES_BY_SKU_QUERY = f"""
    SELECT
        TO_CHAR(DATE, 'YYYY-MM') AS MONTH, CHANNEL, SKU_GROUP,
        L1_CATEGORY, CATEGORY, META1, META2, META3,
        SUM(GROSS_SALES_VALUE) AS GROSS_SALES_VALUE, SUM(QTY) AS QTY,
        SUM(COGS_SOLD) AS COGS_SOLD, SUM(MRP_VALUE) AS MRP_VALUE
    FROM SNITCH_DB.MAPLEMONK.SALES_FOR_AUTO_3
    WHERE DATE >= '{ROLLING_CUTOFF_DATE}'
    GROUP BY TO_CHAR(DATE, 'YYYY-MM'), CHANNEL, SKU_GROUP, L1_CATEGORY, CATEGORY, META1, META2, META3
"""
# NOTE: this now uses the same rolling ROLLING_CUTOFF_DATE as the main sales
# query (computed fresh from today's date every run -- see top of file), not a
# fixed date. That's what keeps this file's size roughly constant over time
# instead of growing forever as more months of history accumulate.

# Returns at the same monthly SKU_GROUP grain as SALES_BY_SKU_QUERY, for the
# Pareto/Quartile "Return" and "Net Value" columns. Joined directly on raw
# SKU_GROUP (NOT the parent-cleaned version used in the main RETURNS_QUERY) --
# RETURNS_DATA and SALES_FOR_AUTO_3 both come from the same order-line SKU
# identifiers, so they should match exactly without needing the meta_mapping
# product-master join at all here. Only OVERALL_RETURNS_QTY exists (no return
# *value* field in the source table) -- return value in Rupee-metric terms is
# therefore an ESTIMATE (return_qty x that SKU-month's average per-unit rate),
# clearly labeled "Est." in the UI. Qty-based return counts are exact.
RETURNS_BY_SKU_QUERY = f"""
    SELECT
        TO_CHAR(DATE, 'YYYY-MM') AS MONTH, SKU_GROUP,
        SUM(OVERALL_RETURNS_QTY) AS RETURN_QTY
    FROM SNITCH_DB.MAPLEMONK.RETURNS_DATA
    WHERE DATE >= '{ROLLING_CUTOFF_DATE}'
    GROUP BY TO_CHAR(DATE, 'YYYY-MM'), SKU_GROUP
"""

# ---------------------------------------------------------------------------
# Snowflake: one representative image URL per sku_group, for the Pareto /
# Quartile tabs (thumbnail next to each ranked SKU group). Picks the first
# IMAGE-type media (by media index) per sku_group across all its variants.
# ---------------------------------------------------------------------------
IMAGES_QUERY = r"""
WITH images AS (
    SELECT
        UPPER(
    TRIM(
        CASE
            WHEN variant.value:sku::STRING ILIKE '4C%'
              OR variant.value:sku::STRING ILIKE 'MP%'
            THEN REGEXP_SUBSTR(variant.value:sku::STRING, '^[^-]+-[^-]+-[^-]+')
            ELSE REGEXP_SUBSTR(variant.value:sku::STRING, '^[^-]+-[^-]+')
        END
    )
) AS sku_group,
        image.value:preview:image:url::STRING AS image_url,
        image.index AS image_index
    FROM snitch_db.maplemonk.new_meafields_product_products_graph_ql t,
         LATERAL FLATTEN(input => PARSE_JSON(t.media)) AS image,
         LATERAL FLATTEN(input => PARSE_JSON(t.variants)) AS variant
    WHERE image.value:mediaContentType::STRING = 'IMAGE'
)
SELECT
    sku_group,
    image_url
FROM images
QUALIFY ROW_NUMBER() OVER (
    PARTITION BY sku_group
    ORDER BY image_index
) = 1
"""


def _load_snowflake_creds():
    """Local dev convenience: if snowflake_credentials.json exists, load its
    keys into os.environ (only if not already set). In GitHub Actions, the
    env vars are set directly by the workflow from repo secrets instead, so
    this file is never needed/present there."""
    if os.path.exists(SNOWFLAKE_CREDS_FILE):
        with open(SNOWFLAKE_CREDS_FILE) as f:
            creds = json.load(f)
        for k, v in creds.items():
            os.environ.setdefault(k, v)


def _snowflake_connect():
    import platform
    platform.libc_ver = lambda *a, **k: ('', '')  # Windows Store python.exe workaround
    import snowflake.connector

    _load_snowflake_creds()
    required = ["SNOWFLAKE_ACCOUNT", "SNOWFLAKE_USER", "SNOWFLAKE_PASSWORD",
                "SNOWFLAKE_WAREHOUSE", "SNOWFLAKE_DATABASE", "SNOWFLAKE_SCHEMA"]
    missing = [k for k in required if not os.environ.get(k)]
    if missing:
        print(f"ERROR: missing Snowflake env vars: {', '.join(missing)}")
        print("Locally: create snowflake_credentials.json (see README/chat). In GitHub")
        print("Actions: set these as repo secrets and pass them into the workflow env.")
        sys.exit(1)

    print("Connecting to Snowflake...")
    return snowflake.connector.connect(
        account=os.environ["SNOWFLAKE_ACCOUNT"],
        user=os.environ["SNOWFLAKE_USER"],
        password=os.environ["SNOWFLAKE_PASSWORD"],
        warehouse=os.environ["SNOWFLAKE_WAREHOUSE"],
        database=os.environ["SNOWFLAKE_DATABASE"],
        schema=os.environ["SNOWFLAKE_SCHEMA"],
    )


def fetch_sales_from_snowflake():
    conn = _snowflake_connect()
    try:
        cur = conn.cursor()
        print("Fetching sales from SALES_FOR_AUTO_3 (aggregated to leaf grain)...")
        cur.execute(SNOWFLAKE_QUERY)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    sales = []
    for row in rows:
        (dt, channel, l1, cat, m1, m2, m3, gross_sales, mrp_value, qty, cogs_sold) = row
        d = dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt)[:10]
        sales.append({
            "date": d, "month": month_key(dt), "channel": channel, "l1": l1, "cat": cat,
            "m1": m1, "m2": m2, "m3": m3,
            "gross_sales": float(gross_sales or 0), "mrp_value": float(mrp_value or 0),
            "qty": float(qty or 0), "cogs_sold": float(cogs_sold or 0),
        })
    print(f"  -> {len(sales)} sales rows from Snowflake")
    return sales


def fetch_sales_by_sku_from_snowflake():
    """Separate, monthly-grain SKU-level dataset for Pareto/Quartile tabs only.
    Kept out of the main sales array (see SALES_BY_SKU_QUERY comment) to avoid
    the data.json size blowup."""
    conn = _snowflake_connect()
    try:
        cur = conn.cursor()
        print("Fetching sales_by_sku (monthly grain) from SALES_FOR_AUTO_3...")
        cur.execute(SALES_BY_SKU_QUERY)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    sales_by_sku = []
    for row in rows:
        (month, channel, sku_group, l1, cat, m1, m2, m3, gross_sales, qty, cogs_sold, mrp_value) = row
        sales_by_sku.append({
            "month": month, "channel": channel, "sku_group": sku_group,
            "l1": l1, "cat": cat, "m1": m1, "m2": m2, "m3": m3,
            "gross_sales": float(gross_sales or 0), "qty": float(qty or 0),
            "cogs_sold": float(cogs_sold or 0), "mrp_value": float(mrp_value or 0),
        })
    print(f"  -> {len(sales_by_sku)} sales_by_sku rows from Snowflake")
    return sales_by_sku


def fetch_returns_by_sku_from_snowflake():
    """Monthly-grain per-SKU returns, for the Pareto/Quartile 'Return' and 'Net
    Value' columns. See RETURNS_BY_SKU_QUERY comment for the value-estimation
    caveat (qty is exact, Rupee value is derived client-side as an estimate)."""
    conn = _snowflake_connect()
    try:
        cur = conn.cursor()
        print("Fetching returns_by_sku (monthly grain) from RETURNS_DATA...")
        cur.execute(RETURNS_BY_SKU_QUERY)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    returns_by_sku = []
    for row in rows:
        month, sku_group, return_qty = row
        if not sku_group:
            continue
        returns_by_sku.append({"month": month, "sku_group": sku_group, "return_qty": float(return_qty or 0)})
    print(f"  -> {len(returns_by_sku)} returns_by_sku rows from Snowflake")
    return returns_by_sku


def fetch_images_from_snowflake():
    """One image URL per sku_group -> dict, for O(1) lookup client-side
    (Pareto/Quartile tab thumbnails). Empty dict (not an error) if the query
    returns nothing, so the tabs still render without images."""
    conn = _snowflake_connect()
    try:
        cur = conn.cursor()
        print("Fetching sku_group -> image_url map...")
        cur.execute(IMAGES_QUERY)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    images = {}
    for sku_group, image_url in rows:
        if sku_group and image_url:
            images[sku_group] = image_url
    print(f"  -> {len(images)} sku_group images from Snowflake")
    return images


# ---------------------------------------------------------------------------
# Snowflake: "new" SKU launch tracker (Live SKU Groups tab). One row per
# SKU_group that's had a "new"-type putaway. Tracks whether it has gone live,
# why it hasn't if not, current warehouse/offline stock, and sales since its
# putaway date. See chat for the full field-by-field breakdown.
# ---------------------------------------------------------------------------
LIVE_SKU_QUERY = r"""
with a as (SELECT
            upper(
                trim(
                    REVERSE(
                        SUBSTRING(
                            REVERSE("Item Type skuCode"),
                            CHARINDEX('-', REVERSE("Item Type skuCode")) + 1,
                            LEN("Item Type skuCode")
                        )
                    )
                )
            ) AS SKU_group,
            LOWER(FINAL_TYPE) type,
            "PUTAWAY_UPDATED"::DATE as date,
            SUM("PUTAWAY_COMPLETED_QUANTITY") AS Qty
        FROM
            snitch_db.maplemonk.putaway_tracking
        WHERE
            "PUTAWAY_UPDATED"::DATE < CURRENT_DATE()
            and LOWER(FINAL_TYPE) like'new%'
        GROUP BY
            1,2,3),
a_agg as (
    SELECT
        SKU_group,
        MAX(date) as date,
        SUM(qty) as qty
    FROM a
    GROUP BY 1
),
meta_map AS (
    SELECT
       UPPER(
            IFF(
                UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE 'MP%'
                OR UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE '4C-%',
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '-.*$', '')
            )
        ) AS sku_group_clean,
       COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LONG TAIL' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'PLUS' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LUXE' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'SNITCH' THEN a.l1_category END),
            MIN(a.l1_category)
        ) AS l1_category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.category END),
            MIN(a.category)
        ) AS category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta1 END),
            MIN(a.meta1)
        ) AS meta1,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta2 END),
            MIN(a.meta2)
        ) AS meta2,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta3 END),
            MIN(a.meta3)
        ) AS meta3,
        max(cogs) as cogs
    FROM snitch_db.maplemonk.meta_mapping_cogs_sku a
    GROUP BY 1
),
lkp_agg AS (
    SELECT
        TRIM(UPPER(SKU_GROUP)) AS sku_group_lkp,
        MAX(FINAL_LIVE_DATE) AS FINAL_LIVE_DATE
    FROM snitch_db.maplemonk.base_product
    GROUP BY 1
),
avl_agg AS (
    SELECT
        TRIM(UPPER(SKU_GROUP)) AS sku_group_avl,
        MAX(STATUS) AS STATUS,
        MAX(SKU_CLASS) AS SKU_CLASS
    FROM snitch_db.maplemonk.availability_master_v2
    GROUP BY 1
),
offline_agg AS (
    SELECT
        TRIM(UPPER(SKU_GROUP)) AS sku_group_off,
        SUM(INVENTORY) AS total_inventory,
        SUM(JIT_QTY) AS total_jit_qty,
        COUNT(DISTINCT MARKETPLACE_MAPPED) AS store_count
    FROM snitch_db.maplemonk.offline_master
    GROUP BY 1
),
total_stores AS (
    SELECT COUNT(DISTINCT MARKETPLACE_MAPPED) AS total_store_count
    FROM snitch_db.maplemonk.offline_master
),
wh_inv AS (
    SELECT
        date,
        TRIM(
            UPPER(
                REVERSE(
                    SUBSTRING(
                        REVERSE("Item SkuCode"),
                        POSITION('-' IN REVERSE("Item SkuCode")) + 1
                    )
                )
            )
        ) AS sku_group,
        SUM(inventory) AS wh_qty
    FROM snitch_db.maplemonk.snitch_final_inventory_wh2
    WHERE facility IN ('SAPL-WH2', 'SAPL-WH1', 'SAPL-NORTH-TAURU')
      AND date = CURRENT_DATE()
    GROUP BY 1, 2
),
today_wh_inv AS (
    SELECT
        sku_group AS sku_group_wh,
        SUM(wh_qty) AS today_wh_qty
    FROM wh_inv
    GROUP BY 1
),
sales_since_live AS (
    SELECT
        TRIM(UPPER(h.SKU_GROUP)) AS sku_group_sales,
        SUM(CASE WHEN h.TYPE = 'Store' THEN h.GROSS_QUANTITY ELSE 0 END) AS units_sold_offline,
        SUM(CASE WHEN h.TYPE = 'Shopify' THEN h.GROSS_QUANTITY ELSE 0 END) AS units_sold_shopify,
        SUM(CASE WHEN h.TYPE = 'Marketplace' THEN h.GROSS_QUANTITY ELSE 0 END) AS units_sold_marketplace,
        SUM(h.GROSS_QUANTITY) AS units_sold_total
    FROM snitch_db.maplemonk.horizontal_sales_categories h
    INNER JOIN a_agg ag
        ON TRIM(UPPER(ag.SKU_GROUP)) = TRIM(UPPER(h.SKU_GROUP))
    WHERE h.DATE::DATE >= ag.date
    GROUP BY 1
),
pipeline_skus AS (
    SELECT DISTINCT CASE
        WHEN subvention IS NULL OR TRIM(subvention) = '' THEN NULL
        ELSE
            SPLIT_PART(
                CASE
                    WHEN LEFT(REPLACE(TRIM(SPLIT_PART(subvention, '/', 1)), ' ', ''), 2) = 'R4'
                    THEN '4' || SUBSTR(REPLACE(TRIM(SPLIT_PART(subvention, '/', 1)), ' ', ''), 3)
                    ELSE REPLACE(TRIM(SPLIT_PART(subvention, '/', 1)), ' ', '')
                END,
                '-', 1
            )
            ||
            CASE
                WHEN SPLIT_PART(
                    CASE
                        WHEN LEFT(REPLACE(TRIM(SPLIT_PART(subvention, '/', 1)), ' ', ''), 2) = 'R4'
                        THEN '4' || SUBSTR(REPLACE(TRIM(SPLIT_PART(subvention, '/', 1)), ' ', ''), 3)
                        ELSE REPLACE(TRIM(SPLIT_PART(subvention, '/', 1)), ' ', '')
                    END,
                    '-', 2
                ) <> ''
                THEN '-' || SPLIT_PART(
                    CASE
                        WHEN LEFT(REPLACE(TRIM(SPLIT_PART(subvention, '/', 1)), ' ', ''), 2) = 'R4'
                        THEN '4' || SUBSTR(REPLACE(TRIM(SPLIT_PART(subvention, '/', 1)), ' ', ''), 3)
                        ELSE REPLACE(TRIM(SPLIT_PART(subvention, '/', 1)), ' ', '')
                    END,
                    '-', 2
                )
                ELSE ''
            END
    END AS normalized_sku
    FROM snitch_db.maplemonk.gs_product_tracking_new_main
    WHERE sku_status_ not in ('Delivered', 'Cancel', 'Hold')
        and LOWER(sku_status_) not LIKE 'deliver%'
        and LOWER(sku_status_) not LIKE 'cancel%'
        and LOWER(sku_status_) not LIKE 'hold'
),
pipeline_variants AS (
    SELECT DISTINCT
        REGEXP_REPLACE(normalized_sku, '-[^-]+$', '') AS parent_sku,
        UPPER(TRIM(normalized_sku)) AS variant_key
    FROM pipeline_skus
    WHERE normalized_sku IS NOT NULL
),
logic_variants AS (
    SELECT DISTINCT
        UPPER(TRIM(SKU_MAPPING)) AS parent_sku,
        UPPER(TRIM(SKU_MAPPING)) || '-' || UPPER(TRIM(COLOR)) AS variant_key
    FROM snitch_db.maplemonk.logic_final_item_master
    WHERE SKU_MAPPING IS NOT NULL AND COLOR IS NOT NULL
),
expected_colors AS (
    SELECT parent_sku, COUNT(DISTINCT variant_key) AS expected_colors
    FROM (
        SELECT parent_sku, variant_key FROM pipeline_variants
        UNION
        SELECT parent_sku, variant_key FROM logic_variants
    )
    GROUP BY 1
),
actual_colors AS (
    SELECT
        REGEXP_REPLACE(SKU_group, '-[^-]+$', '') AS parent_sku,
        COUNT(DISTINCT SKU_group) AS actual_colors
    FROM a_agg
    GROUP BY 1
),
capsule_map AS (
    SELECT DISTINCT TRIM(UPPER(SKU)) AS sku_group_capsule
    FROM snitch_db.maplemonk.collection_sku_mapped_sheet1
),
delivery_status_map AS (
    SELECT
        TRIM(UPPER(SKU_GROUP)) AS sku_group_delivery,
        MAX(DELIVERY_STATUS_NORTH) AS delivery_status_north,
        MAX(DELIVERY_STATUS_SOUTH) AS delivery_status_south,
        MAX(NORTH_PO) AS north_po,
        MAX(NORTH_INWARDED) AS north_inwarded,
        MAX(SOUTH_PO) AS south_po,
        MAX(SOUTH_INWARDED) AS south_inwarded
    FROM snitch_db.maplemonk.SKU_LIVE_MECH_V2
    GROUP BY 1
)
select
    a_agg.date,
    a_agg.sku_group,
    meta_map.l1_category,
    meta_map.category,
    meta_map.meta1,
    meta_map.meta2,
    meta_map.meta3,
    lkp_agg.FINAL_LIVE_DATE,
    avl_agg.STATUS,
    avl_agg.SKU_CLASS,
    a_agg.qty*meta_map.cogs as cogs_value,
    a_agg.qty as qty,
    DATEDIFF(day, a_agg.date, lkp_agg.FINAL_LIVE_DATE) as delay,
    offline_agg.total_inventory,
    offline_agg.total_jit_qty,
    offline_agg.store_count,
    total_stores.total_store_count,
    total_stores.total_store_count - COALESCE(offline_agg.store_count, 0) AS stores_not_allocated,
    today_wh_inv.today_wh_qty,
    sales_since_live.units_sold_offline,
    sales_since_live.units_sold_shopify,
    sales_since_live.units_sold_marketplace,
    sales_since_live.units_sold_total,
    CASE WHEN capsule_map.sku_group_capsule IS NOT NULL THEN 'CAPSULE' END AS capsule_tag,
    CASE
        WHEN lkp_agg.FINAL_LIVE_DATE IS NULL
         AND avl_agg.SKU_CLASS <> 'MP'
         AND actual_colors.actual_colors < expected_colors.expected_colors
        THEN 'All colours not inwarded'

        WHEN lkp_agg.FINAL_LIVE_DATE IS NULL
         AND avl_agg.SKU_CLASS <> 'MP'
         AND (
                UPPER(TRIM(delivery_status_map.delivery_status_north)) LIKE '%NOT_YET%'
             OR (delivery_status_map.north_po > 0 AND COALESCE(delivery_status_map.north_inwarded, 0) < delivery_status_map.north_po)
             )
         AND (
                UPPER(TRIM(delivery_status_map.delivery_status_south)) LIKE '%NOT_YET%'
             OR (delivery_status_map.south_po > 0 AND COALESCE(delivery_status_map.south_inwarded, 0) < delivery_status_map.south_po)
             )
        THEN 'Inward pending in both WH'

        WHEN lkp_agg.FINAL_LIVE_DATE IS NULL
         AND avl_agg.SKU_CLASS <> 'MP'
         AND (
                UPPER(TRIM(delivery_status_map.delivery_status_north)) LIKE '%NOT_YET%'
             OR (delivery_status_map.north_po > 0 AND COALESCE(delivery_status_map.north_inwarded, 0) < delivery_status_map.north_po)
             )
        THEN 'Delivery pending in North'

        WHEN lkp_agg.FINAL_LIVE_DATE IS NULL
         AND avl_agg.SKU_CLASS <> 'MP'
         AND (
                UPPER(TRIM(delivery_status_map.delivery_status_south)) LIKE '%NOT_YET%'
             OR (delivery_status_map.south_po > 0 AND COALESCE(delivery_status_map.south_inwarded, 0) < delivery_status_map.south_po)
             )
        THEN 'Delivery pending in South'
    END AS not_live_remark
from a_agg
cross join total_stores
left join meta_map
    on meta_map.sku_group_clean = UPPER(
        IFF(
            UPPER(REPLACE(a_agg.SKU_GROUP, ' ', '')) LIKE 'MP%'
            OR UPPER(REPLACE(a_agg.SKU_GROUP, ' ', '')) LIKE '4C-%',
            REGEXP_REPLACE(REPLACE(a_agg.SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
            REGEXP_REPLACE(REPLACE(a_agg.SKU_GROUP, ' ', ''), '-.*$', '')
        )
    )
left join lkp_agg
    on lkp_agg.sku_group_lkp = TRIM(UPPER(a_agg.SKU_GROUP))
left join avl_agg
    on avl_agg.sku_group_avl = TRIM(UPPER(a_agg.SKU_GROUP))
left join offline_agg
    on offline_agg.sku_group_off = TRIM(UPPER(a_agg.SKU_GROUP))
left join today_wh_inv
    on today_wh_inv.sku_group_wh = TRIM(UPPER(a_agg.SKU_GROUP))
left join sales_since_live
    on sales_since_live.sku_group_sales = TRIM(UPPER(a_agg.SKU_GROUP))
left join expected_colors
    on expected_colors.parent_sku = REGEXP_REPLACE(a_agg.SKU_group, '-[^-]+$', '')
left join actual_colors
    on actual_colors.parent_sku = REGEXP_REPLACE(a_agg.SKU_group, '-[^-]+$', '')
left join capsule_map
    on capsule_map.sku_group_capsule = TRIM(UPPER(a_agg.SKU_GROUP))
left join delivery_status_map
    on delivery_status_map.sku_group_delivery = TRIM(UPPER(a_agg.SKU_GROUP))
"""


def fetch_live_sku_groups_from_snowflake():
    conn = _snowflake_connect()
    try:
        cur = conn.cursor()
        print("Fetching Live SKU Groups (new-launch tracker) from putaway_tracking...")
        cur.execute(LIVE_SKU_QUERY)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    def _d(v):
        if v is None:
            return None
        return v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v)[:10]

    live_sku_groups = []
    for row in rows:
        (date, sku_group, l1, cat, m1, m2, m3, final_live_date, status, sku_class,
         cogs_value, qty, delay, total_inventory, total_jit_qty, store_count,
         total_store_count, stores_not_allocated, today_wh_qty,
         units_sold_offline, units_sold_shopify, units_sold_marketplace, units_sold_total,
         capsule_tag, not_live_remark) = row
        live_sku_groups.append({
            "date": _d(date), "sku_group": sku_group, "l1": l1, "cat": cat, "m1": m1, "m2": m2, "m3": m3,
            "live_date": _d(final_live_date), "status": status, "sku_class": sku_class,
            "cogs_value": float(cogs_value or 0), "qty": float(qty or 0), "delay": int(delay) if delay is not None else None,
            "total_inventory": float(total_inventory or 0), "total_jit_qty": float(total_jit_qty or 0),
            "store_count": int(store_count or 0), "total_store_count": int(total_store_count or 0),
            "stores_not_allocated": int(stores_not_allocated or 0), "today_wh_qty": float(today_wh_qty or 0),
            "units_sold_offline": float(units_sold_offline or 0), "units_sold_shopify": float(units_sold_shopify or 0),
            "units_sold_marketplace": float(units_sold_marketplace or 0), "units_sold_total": float(units_sold_total or 0),
            "capsule_tag": capsule_tag, "not_live_remark": not_live_remark,
        })
    print(f"  -> {len(live_sku_groups)} live SKU group rows from Snowflake")
    return live_sku_groups


# ---------------------------------------------------------------------------
# Snowflake: post-launch click performance (Clicks tab). One row per live
# SKU_group (FINAL_LIVE_DATE required, MP/FK/B2B/PP prefixes excluded at the
# source), with clicks in its first 15 days live, next 15 days, and the
# trailing 30 days -- alongside current warehouse stock, for spotting
# well-stocked-but-cold SKUs vs. thin-stocked-but-hot ones.
# ---------------------------------------------------------------------------
CLICKS_QUERY = r"""
with a as (
    SELECT
        upper(trim(REVERSE(SUBSTRING(REVERSE("Item Type skuCode"),CHARINDEX('-', REVERSE("Item Type skuCode")) + 1,LEN("Item Type skuCode"))))) AS SKU_group,
        LOWER(FINAL_TYPE) type,
        "PUTAWAY_UPDATED"::DATE as date,
        SUM("PUTAWAY_COMPLETED_QUANTITY") AS Qty
    FROM snitch_db.maplemonk.putaway_tracking
    WHERE "PUTAWAY_UPDATED"::DATE < CURRENT_DATE()
      and LOWER(FINAL_TYPE) like 'new%'
    GROUP BY 1,2,3
),
a_agg as (
    SELECT SKU_group, MAX(date) as date, SUM(qty) as qty
    FROM a
    GROUP BY 1
),
meta_map AS (
    SELECT
        UPPER(
            IFF(
                UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE 'MP%'
                OR UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE '4C-%',
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '-.*$', '')
            )
        ) AS sku_group_clean,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LONG TAIL' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'PLUS' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LUXE' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'SNITCH' THEN a.l1_category END),
            MIN(a.l1_category)
        ) AS l1_category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.category END),
            MIN(a.category)
        ) AS category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta1 END),
            MIN(a.meta1)
        ) AS meta1,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta2 END),
            MIN(a.meta2)
        ) AS meta2,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta3 END),
            MIN(a.meta3)
        ) AS meta3
    FROM snitch_db.maplemonk.meta_mapping_cogs_sku a
    GROUP BY 1
),
lkp_agg AS (
    SELECT
        TRIM(UPPER(SKU_GROUP)) AS sku_group_lkp,
        MAX(FINAL_LIVE_DATE) AS FINAL_LIVE_DATE
    FROM snitch_db.maplemonk.base_product
    GROUP BY 1
),
wh_inv AS (
    SELECT
        date,
        TRIM(UPPER(REVERSE(SUBSTRING(REVERSE("Item SkuCode"), POSITION('-' IN REVERSE("Item SkuCode")) + 1)))) AS sku_group,
        SPLIT_PART("Item SkuCode", '-', -1) AS size_raw,
        SUM(inventory) AS wh_qty
    FROM snitch_db.maplemonk.snitch_final_inventory_wh2
    WHERE facility IN ('SAPL-WH2', 'SAPL-WH1', 'SAPL-NORTH-TAURU')
      AND date = CURRENT_DATE()
    GROUP BY 1, 2, 3
),
today_wh_inv AS (
    SELECT sku_group AS sku_group_wh, SUM(wh_qty) AS today_wh_qty
    FROM wh_inv
    GROUP BY 1
),
size_agg AS (
    SELECT
        sku_group AS sku_group_size,
        CASE
            WHEN UPPER(TRIM(size_raw)) IN ('XS','S','M','L','XL','XXL','3XL','4XL','5XL','6XL') THEN UPPER(TRIM(size_raw))
            WHEN TRIM(size_raw) = '28' THEN 'XS'
            WHEN TRIM(size_raw) = '30' THEN 'S'
            WHEN TRIM(size_raw) = '32' THEN 'M'
            WHEN TRIM(size_raw) = '34' THEN 'L'
            WHEN TRIM(size_raw) = '36' THEN 'XL'
            WHEN TRIM(size_raw) = '38' THEN 'XXL'
            WHEN TRIM(size_raw) = '40' THEN '3XL'
            WHEN TRIM(size_raw) = '42' THEN '4XL'
            WHEN TRIM(size_raw) = '44' THEN '5XL'
            WHEN TRIM(size_raw) = '46' THEN '6XL'
            ELSE UPPER(TRIM(size_raw))
        END AS size_mapped,
        SUM(wh_qty) AS wh_qty
    FROM wh_inv
    GROUP BY 1, 2
),
size_pivot AS (
    SELECT
        sku_group_size,
        SUM(CASE WHEN size_mapped = 'XS' THEN wh_qty ELSE 0 END) AS wh_xs,
        SUM(CASE WHEN size_mapped = 'S' THEN wh_qty ELSE 0 END) AS wh_s,
        SUM(CASE WHEN size_mapped = 'M' THEN wh_qty ELSE 0 END) AS wh_m,
        SUM(CASE WHEN size_mapped = 'L' THEN wh_qty ELSE 0 END) AS wh_l,
        SUM(CASE WHEN size_mapped = 'XL' THEN wh_qty ELSE 0 END) AS wh_xl,
        SUM(CASE WHEN size_mapped = 'XXL' THEN wh_qty ELSE 0 END) AS wh_xxl,
        SUM(CASE WHEN size_mapped = '3XL' THEN wh_qty ELSE 0 END) AS wh_3xl,
        SUM(CASE WHEN size_mapped = '4XL' THEN wh_qty ELSE 0 END) AS wh_4xl,
        SUM(CASE WHEN size_mapped = '5XL' THEN wh_qty ELSE 0 END) AS wh_5xl,
        SUM(CASE WHEN size_mapped = '6XL' THEN wh_qty ELSE 0 END) AS wh_6xl
    FROM size_agg
    GROUP BY 1
),
offline_agg AS (
    SELECT
        TRIM(UPPER(SKU_GROUP)) AS sku_group_off,
        COUNT(DISTINCT CASE WHEN INVENTORY > 0 OR JIT_QTY > 0 THEN MARKETPLACE_MAPPED END) AS store_count,
        SUM(INVENTORY) AS total_inventory,
        SUM(JIT_QTY) AS total_jit_qty
    FROM snitch_db.maplemonk.offline_master
    GROUP BY 1
),
clicks_agg AS (
    SELECT
        TRIM(UPPER(SKU_GROUP)) AS sku_group_clicks,
        ga_date::DATE AS click_date,
        SUM(clicks) AS clicks
    FROM snitch_db.maplemonk.clicks_itemid
    GROUP BY 1, 2
),
image_agg AS (
    SELECT
        TRIM(UPPER(sku_group)) AS sku_group_image,
        MAX(image_url) AS image_url
    FROM (
        SELECT
            LEFT(variant.value:sku::STRING, LENGTH(variant.value:sku::STRING) - POSITION('-' IN REVERSE(variant.value:sku::STRING))) AS sku_group,
            image.value:"preview":"image":"url"::string AS image_url
        FROM
            snitch_db.maplemonk.new_meafields_product_products_graph_ql t,
            LATERAL FLATTEN(input => parse_json(t.media)) AS image,
            LATERAL FLATTEN(input => PARSE_JSON(t.variants)) AS variant
        WHERE
            image.value:"mediaContentType"::string = 'IMAGE'
        QUALIFY ROW_NUMBER() OVER (PARTITION BY t.id ORDER BY image.index) = 1
    )
    GROUP BY 1
),
sales_window AS (
    SELECT
        lk.sku_group_lkp AS sku_group_sw,
        GREATEST(lk.FINAL_LIVE_DATE, DATEADD(day, -30, CURRENT_DATE())) AS window_start,
        DATEDIFF(day, GREATEST(lk.FINAL_LIVE_DATE, DATEADD(day, -30, CURRENT_DATE())), CURRENT_DATE()) + 1 AS window_days
    FROM lkp_agg lk
    WHERE lk.FINAL_LIVE_DATE IS NOT NULL
),
sales_30 AS (
    SELECT
        TRIM(UPPER(h.SKU_GROUP)) AS sku_group_sales,
        SUM(CASE WHEN h.TYPE = 'Store' THEN h.GROSS_QUANTITY ELSE 0 END) AS retail_sales,
        SUM(CASE WHEN h.TYPE = 'Shopify' THEN h.GROSS_QUANTITY ELSE 0 END) AS shopify_sales,
        SUM(CASE WHEN h.TYPE = 'Marketplace' THEN h.GROSS_QUANTITY ELSE 0 END) AS mp_sales,
        SUM(h.GROSS_QUANTITY) AS ttl_sales,
        MAX(sw.window_days) AS window_days
    FROM snitch_db.maplemonk.horizontal_sales_categories h
    INNER JOIN sales_window sw
        ON sw.sku_group_sw = TRIM(UPPER(h.SKU_GROUP))
    WHERE h.DATE::DATE >= sw.window_start
      AND h.DATE::DATE <= CURRENT_DATE()
    GROUP BY 1
)
select
    a_agg.SKU_group,                                              -- 1
    lkp_agg.FINAL_LIVE_DATE,                                       -- 2
    meta_map.l1_category,                                          -- 3
    meta_map.category,                                             -- 4
    meta_map.meta1,                                                -- 5
    meta_map.meta2,                                                -- 6
    meta_map.meta3,                                                -- 7
    today_wh_inv.today_wh_qty AS total_wh_inventory,                -- 8
    size_pivot.wh_xs,                                              -- 9
    size_pivot.wh_s,                                               -- 10
    size_pivot.wh_m,                                               -- 11
    size_pivot.wh_l,                                               -- 12
    size_pivot.wh_xl,                                              -- 13
    size_pivot.wh_xxl,                                             -- 14
    size_pivot.wh_3xl,                                             -- 15
    size_pivot.wh_4xl,                                             -- 16
    size_pivot.wh_5xl,                                             -- 17
    size_pivot.wh_6xl,                                             -- 18
    CASE
        WHEN UPPER(TRIM(meta_map.l1_category)) IN ('SNITCH','LUXE') THEN
            LEAST(
                FLOOR(COALESCE(size_pivot.wh_s,0) / 1),
                FLOOR(COALESCE(size_pivot.wh_m,0) / 2),
                FLOOR(COALESCE(size_pivot.wh_l,0) / 2),
                FLOOR(COALESCE(size_pivot.wh_xl,0) / 1)
            )
        WHEN UPPER(TRIM(meta_map.l1_category)) = 'PLUS' THEN
            LEAST(
                FLOOR(COALESCE(size_pivot.wh_3xl,0) / 1),
                FLOOR(COALESCE(size_pivot.wh_4xl,0) / 1),
                FLOOR(COALESCE(size_pivot.wh_5xl,0) / 1),
                FLOOR(COALESCE(size_pivot.wh_6xl,0) / 1)
            )
        ELSE NULL
    END AS sets_available,                                         -- 19
    offline_agg.store_count,                                       -- 20
    COALESCE(offline_agg.total_inventory,0) + COALESCE(offline_agg.total_jit_qty,0) AS total_store_inventory, -- 21
    SUM(CASE WHEN c.click_date BETWEEN lkp_agg.FINAL_LIVE_DATE AND DATEADD(day, 14, lkp_agg.FINAL_LIVE_DATE) THEN c.clicks ELSE 0 END) AS clicks_day_1_15,   -- 22
    SUM(CASE WHEN c.click_date BETWEEN DATEADD(day, 15, lkp_agg.FINAL_LIVE_DATE) AND DATEADD(day, 29, lkp_agg.FINAL_LIVE_DATE) THEN c.clicks ELSE 0 END) AS clicks_day_15_30, -- 23
    SUM(CASE WHEN c.click_date BETWEEN DATEADD(day, -30, CURRENT_DATE()) AND CURRENT_DATE() THEN c.clicks ELSE 0 END) AS clicks_last_30_days,                -- 24
    sales_30.ttl_sales,                                            -- 25
    sales_30.retail_sales,                                         -- 26
    sales_30.shopify_sales,                                        -- 27
    sales_30.mp_sales,                                             -- 28
    ROUND(
        (COALESCE(today_wh_inv.today_wh_qty,0) + COALESCE(offline_agg.total_inventory,0) + COALESCE(offline_agg.total_jit_qty,0))
        / NULLIF(sales_30.ttl_sales / NULLIF(sales_30.window_days,0), 0)
    , 1) AS overall_doi,                                           -- 29
    ROUND(
        (COALESCE(offline_agg.total_inventory,0) + COALESCE(offline_agg.total_jit_qty,0))
        / NULLIF(sales_30.retail_sales / NULLIF(sales_30.window_days,0), 0)
    , 1) AS retail_doi,                                            -- 30
    ROUND(
        COALESCE(today_wh_inv.today_wh_qty,0)
        / NULLIF(sales_30.shopify_sales / NULLIF(sales_30.window_days,0), 0)
    , 1) AS shopify_doi,                                           -- 31
    ROUND(
        COALESCE(today_wh_inv.today_wh_qty,0)
        / NULLIF(sales_30.mp_sales / NULLIF(sales_30.window_days,0), 0)
    , 1) AS mp_doi,                                                -- 32
    image_agg.image_url                                            -- 33
from a_agg
left join meta_map
    on meta_map.sku_group_clean = UPPER(
        IFF(
            UPPER(REPLACE(a_agg.SKU_GROUP, ' ', '')) LIKE 'MP%'
            OR UPPER(REPLACE(a_agg.SKU_GROUP, ' ', '')) LIKE '4C-%',
            REGEXP_REPLACE(REPLACE(a_agg.SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
            REGEXP_REPLACE(REPLACE(a_agg.SKU_GROUP, ' ', ''), '-.*$', '')
        )
    )
inner join lkp_agg
    on lkp_agg.sku_group_lkp = TRIM(UPPER(a_agg.SKU_GROUP))
    and lkp_agg.FINAL_LIVE_DATE IS NOT NULL
left join today_wh_inv
    on today_wh_inv.sku_group_wh = TRIM(UPPER(a_agg.SKU_GROUP))
left join size_pivot
    on size_pivot.sku_group_size = TRIM(UPPER(a_agg.SKU_GROUP))
left join offline_agg
    on offline_agg.sku_group_off = TRIM(UPPER(a_agg.SKU_GROUP))
left join clicks_agg c
    on c.sku_group_clicks = TRIM(UPPER(a_agg.SKU_GROUP))
left join image_agg
    on image_agg.sku_group_image = TRIM(UPPER(a_agg.SKU_GROUP))
left join sales_30
    on sales_30.sku_group_sales = TRIM(UPPER(a_agg.SKU_GROUP))
where
    UPPER(a_agg.SKU_GROUP) NOT LIKE 'MP%'
    AND UPPER(a_agg.SKU_GROUP) NOT LIKE 'FK%'
    AND UPPER(a_agg.SKU_GROUP) NOT LIKE 'B2B%'
    AND UPPER(a_agg.SKU_GROUP) NOT LIKE 'PP%'
group by
    a_agg.SKU_group,
    lkp_agg.FINAL_LIVE_DATE,
    meta_map.l1_category,
    meta_map.category,
    meta_map.meta1,
    meta_map.meta2,
    meta_map.meta3,
    today_wh_inv.today_wh_qty,
    size_pivot.wh_xs,
    size_pivot.wh_s,
    size_pivot.wh_m,
    size_pivot.wh_l,
    size_pivot.wh_xl,
    size_pivot.wh_xxl,
    size_pivot.wh_3xl,
    size_pivot.wh_4xl,
    size_pivot.wh_5xl,
    size_pivot.wh_6xl,
    offline_agg.store_count,
    offline_agg.total_inventory,
    offline_agg.total_jit_qty,
    sales_30.ttl_sales,
    sales_30.retail_sales,
    sales_30.shopify_sales,
    sales_30.mp_sales,
    sales_30.window_days,
    image_agg.image_url
order by
    total_wh_inventory DESC,
    clicks_last_30_days DESC
"""


def fetch_clicks_from_snowflake():
    conn = _snowflake_connect()
    try:
        cur = conn.cursor()
        print("Fetching Clicks (post-launch click performance) from clicks_itemid...")
        cur.execute(CLICKS_QUERY)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    def _d(v):
        if v is None:
            return None
        return v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v)[:10]

    def _f(v):
        return float(v) if v is not None else None

    clicks = []
    for row in rows:
        (sku_group, final_live_date, l1, cat, m1, m2, m3, total_wh_inventory,
         wh_xs, wh_s, wh_m, wh_l, wh_xl, wh_xxl, wh_3xl, wh_4xl, wh_5xl, wh_6xl,
         sets_available, store_count, total_store_inventory,
         clicks_1_15, clicks_15_30, clicks_last_30,
         ttl_sales, retail_sales, shopify_sales, mp_sales,
         overall_doi, retail_doi, shopify_doi, mp_doi,
         image_url) = row
        clicks.append({
            "sku_group": sku_group, "live_date": _d(final_live_date), "l1": l1, "cat": cat,
            "m1": m1, "m2": m2, "m3": m3, "wh_qty": float(total_wh_inventory or 0),
            # Warehouse stock by size -- feeds "sets available" (matched sizes across
            # a category's core size run, not just raw total units).
            "wh_xs": float(wh_xs or 0), "wh_s": float(wh_s or 0), "wh_m": float(wh_m or 0),
            "wh_l": float(wh_l or 0), "wh_xl": float(wh_xl or 0), "wh_xxl": float(wh_xxl or 0),
            "wh_3xl": float(wh_3xl or 0), "wh_4xl": float(wh_4xl or 0), "wh_5xl": float(wh_5xl or 0),
            "wh_6xl": float(wh_6xl or 0),
            # None (not 0) when the L1 isn't SNITCH/LUXE/PLUS -- "not applicable",
            # distinct from "zero sets available".
            "sets_available": _f(sets_available),
            "store_count": int(store_count or 0), "total_store_inventory": float(total_store_inventory or 0),
            "clicks_1_15": float(clicks_1_15 or 0), "clicks_15_30": float(clicks_15_30 or 0),
            "clicks_last_30": float(clicks_last_30 or 0),
            # Trailing-30-day (or since-live, if younger) sales by channel, and the
            # resulting DOI figures. All four DOI fields stay None (not 0/inf) when
            # there's no sales in the window to divide by.
            "ttl_sales": float(ttl_sales or 0), "retail_sales": float(retail_sales or 0),
            "shopify_sales": float(shopify_sales or 0), "mp_sales": float(mp_sales or 0),
            "overall_doi": _f(overall_doi), "retail_doi": _f(retail_doi),
            "shopify_doi": _f(shopify_doi), "mp_doi": _f(mp_doi),
            "image_url": image_url,
        })
    print(f"  -> {len(clicks)} clicks rows from Snowflake")
    return clicks


# Separate monthly-grain click totals (last 6 calendar months, all SKU groups
# with any clicks, not just live ones) for the Clicks tab's MoM trend chart.
# Kept as its own query/output key rather than folded into CLICKS_QUERY since
# it answers a different question (overall click volume trend) at a
# different grain (month, not launch-relative windows).
CLICKS_MONTHLY_QUERY = r"""
WITH meta_map AS (
    SELECT
        UPPER(
            IFF(
                UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE 'MP%'
                OR UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE '4C-%',
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '-.*$', '')
            )
        ) AS sku_group_clean,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LONG TAIL' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'PLUS' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LUXE' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'SNITCH' THEN a.l1_category END),
            MIN(a.l1_category)
        ) AS l1_category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.category END),
            MIN(a.category)
        ) AS category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta1 END),
            MIN(a.meta1)
        ) AS meta1,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta2 END),
            MIN(a.meta2)
        ) AS meta2,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta3 END),
            MIN(a.meta3)
        ) AS meta3
    FROM snitch_db.maplemonk.meta_mapping_cogs_sku a
    GROUP BY 1
),
agg AS (
    SELECT
        TRIM(UPPER(SKU_GROUP)) AS sku_group_clicks,
        DATE_TRUNC('month', ga_date)::DATE AS month,
        SUM(clicks) AS clicks
    FROM snitch_db.maplemonk.clicks_itemid
    WHERE ga_date >= DATEADD(month, -6, DATE_TRUNC('month', CURRENT_DATE()))
    GROUP BY 1, 2
)
SELECT
    agg.sku_group_clicks,
    TO_CHAR(agg.month, 'YYYY-MM') AS month,
    agg.clicks,
    meta_map.l1_category,
    meta_map.category,
    meta_map.meta1,
    meta_map.meta2,
    meta_map.meta3
FROM agg
LEFT JOIN meta_map
    ON meta_map.sku_group_clean = UPPER(
        IFF(
            UPPER(REPLACE(agg.sku_group_clicks, ' ', '')) LIKE 'MP%'
            OR UPPER(REPLACE(agg.sku_group_clicks, ' ', '')) LIKE '4C-%',
            REGEXP_REPLACE(REPLACE(agg.sku_group_clicks, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
            REGEXP_REPLACE(REPLACE(agg.sku_group_clicks, ' ', ''), '-.*$', '')
        )
    )
WHERE
    UPPER(agg.sku_group_clicks) NOT LIKE 'MP%'
    AND UPPER(agg.sku_group_clicks) NOT LIKE 'FK%'
    AND UPPER(agg.sku_group_clicks) NOT LIKE 'B2B%'
    AND UPPER(agg.sku_group_clicks) NOT LIKE 'PP%'
"""


def fetch_clicks_monthly_from_snowflake():
    conn = _snowflake_connect()
    try:
        cur = conn.cursor()
        print("Fetching Clicks Monthly (last 6 months MoM trend) from clicks_itemid...")
        cur.execute(CLICKS_MONTHLY_QUERY)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    clicks_monthly = []
    for row in rows:
        sku_group, month, clicks_val, l1, cat, m1, m2, m3 = row
        clicks_monthly.append({
            "sku_group": sku_group, "month": month, "clicks": float(clicks_val or 0),
            "l1": l1, "cat": cat, "m1": m1, "m2": m2, "m3": m3,
        })
    print(f"  -> {len(clicks_monthly)} clicks_monthly rows from Snowflake")
    return clicks_monthly


# ---------------------------------------------------------------------------
# Snowflake: Store Cut Size (Store Cut Size tab). Per L1 x Category x Meta1-3
# x Cluster (store region, plus a synthetic 'OVERALL' row per combo), what
# share of stores carry a "cut" size run (missing a core size) vs "full" for
# that product line, as of the latest daily offline snapshot.
# ---------------------------------------------------------------------------
STORE_CUT_SIZE_QUERY = r"""
WITH meta_map AS (
    SELECT
        UPPER(
            IFF(
                UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE 'MP%'
                OR UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE '4C-%',
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '-.*$', '')
            )
        ) AS sku_group_clean,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LONG TAIL' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'PLUS' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LUXE' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'SNITCH' THEN a.l1_category END),
            MIN(a.l1_category)
        ) AS l1_category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.category END),
            MIN(a.category)
        ) AS category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta1 END),
            MIN(a.meta1)
        ) AS meta1,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta2 END),
            MIN(a.meta2)
        ) AS meta2,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta3 END),
            MIN(a.meta3)
        ) AS meta3
    FROM snitch_db.maplemonk.meta_mapping_cogs_sku a
    GROUP BY 1
),
latest_date AS (
    SELECT MAX("DATE"::DATE) AS max_date
    FROM snitch_db.maplemonk.offline_master_daily_report_1
),
base AS (
    SELECT
        om.SKU_GROUP,
        om.BRANCH_CODE,
        mm.l1_category,
        mm.category,
        mm.meta1,
        mm.meta2,
        mm.meta3,
        om.CLUSTER,
        COALESCE(om.INVENTORY,0) + COALESCE(om.JIT_QTY,0) AS TTL_INV,
        CASE
            WHEN UPPER(TRIM(mm.l1_category)) IN ('SNITCH','LUXE') THEN
                CASE WHEN
                    (COALESCE(om.S_UNITS,0) + COALESCE(om.JIT_S_UNITS,0) > 0
                     AND COALESCE(om.M_UNITS,0) + COALESCE(om.JIT_M_UNITS,0) > 0
                     AND COALESCE(om.L_UNITS,0) + COALESCE(om.JIT_L_UNITS,0) > 0)
                    OR
                    (COALESCE(om.M_UNITS,0) + COALESCE(om.JIT_M_UNITS,0) > 0
                     AND COALESCE(om.L_UNITS,0) + COALESCE(om.JIT_L_UNITS,0) > 0
                     AND COALESCE(om.XL_UNITS,0) + COALESCE(om.JIT_XL_UNITS,0) > 0)
                THEN 'FULL' ELSE 'CUT' END
            WHEN UPPER(TRIM(mm.l1_category)) = 'PLUS' THEN
                CASE WHEN
                    COALESCE(om.XL3_UNITS,0) + COALESCE(om.JIT_3XL_UNITS,0) > 0
                    AND COALESCE(om.XL4_UNITS,0) + COALESCE(om.JIT_4XL_UNITS,0) > 0
                THEN 'FULL' ELSE 'CUT' END
            ELSE NULL
        END AS SIZE_STATUS
    FROM snitch_db.maplemonk.offline_master_daily_report_1 om
    INNER JOIN latest_date ld
        ON om."DATE"::DATE = ld.max_date
    LEFT JOIN meta_map mm
        ON mm.sku_group_clean = UPPER(
            IFF(
                UPPER(REPLACE(om.SKU_GROUP, ' ', '')) LIKE 'MP%'
                OR UPPER(REPLACE(om.SKU_GROUP, ' ', '')) LIKE '4C-%',
                REGEXP_REPLACE(REPLACE(om.SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
                REGEXP_REPLACE(REPLACE(om.SKU_GROUP, ' ', ''), '-.*$', '')
            )
        )
    WHERE UPPER(TRIM(mm.l1_category)) NOT IN ('LONG TAIL')
        AND mm.l1_category IS NOT NULL
        AND om.CLUSTER IS NOT NULL
),
cluster_store_count AS (
    SELECT
        CLUSTER,
        COUNT(DISTINCT BRANCH_CODE) AS total_store_count_in_cluster
    FROM base
    GROUP BY 1
),
overall_store_count AS (
    SELECT
        COUNT(DISTINCT BRANCH_CODE) AS total_store_count_overall
    FROM base
)
SELECT
    b.l1_category,
    b.category,
    b.meta1,
    b.meta2,
    b.meta3,
    b.CLUSTER,
    csc.total_store_count_in_cluster,
    COUNT(DISTINCT CASE WHEN b.TTL_INV > 0 THEN b.SKU_GROUP END) AS unique_option_count,
    COUNT(DISTINCT CASE WHEN b.TTL_INV > 0 AND b.SIZE_STATUS = 'CUT' THEN b.BRANCH_CODE END) AS store_count_with_cut_size,
    ROUND(
        100.0 * COUNT(CASE WHEN b.TTL_INV > 0 AND b.SIZE_STATUS = 'CUT' THEN 1 END)
        / NULLIF(COUNT(CASE WHEN b.TTL_INV > 0 THEN 1 END), 0)
    , 1) AS pct_cut,
    ROUND(
        100.0 * COUNT(CASE WHEN b.TTL_INV > 0 AND b.SIZE_STATUS = 'FULL' THEN 1 END)
        / NULLIF(COUNT(CASE WHEN b.TTL_INV > 0 THEN 1 END), 0)
    , 1) AS pct_full
FROM base b
LEFT JOIN cluster_store_count csc
    ON csc.CLUSTER = b.CLUSTER
GROUP BY
    b.l1_category,
    b.category,
    b.meta1,
    b.meta2,
    b.meta3,
    b.CLUSTER,
    csc.total_store_count_in_cluster

UNION ALL

SELECT
    b.l1_category,
    b.category,
    b.meta1,
    b.meta2,
    b.meta3,
    'OVERALL' AS CLUSTER,
    osc.total_store_count_overall,
    COUNT(DISTINCT CASE WHEN b.TTL_INV > 0 THEN b.SKU_GROUP END) AS unique_option_count,
    COUNT(DISTINCT CASE WHEN b.TTL_INV > 0 AND b.SIZE_STATUS = 'CUT' THEN b.BRANCH_CODE END) AS store_count_with_cut_size,
    ROUND(
        100.0 * COUNT(CASE WHEN b.TTL_INV > 0 AND b.SIZE_STATUS = 'CUT' THEN 1 END)
        / NULLIF(COUNT(CASE WHEN b.TTL_INV > 0 THEN 1 END), 0)
    , 1) AS pct_cut,
    ROUND(
        100.0 * COUNT(CASE WHEN b.TTL_INV > 0 AND b.SIZE_STATUS = 'FULL' THEN 1 END)
        / NULLIF(COUNT(CASE WHEN b.TTL_INV > 0 THEN 1 END), 0)
    , 1) AS pct_full
FROM base b
CROSS JOIN overall_store_count osc
GROUP BY
    b.l1_category,
    b.category,
    b.meta1,
    b.meta2,
    b.meta3,
    osc.total_store_count_overall

ORDER BY
    l1_category,
    category,
    meta1,
    meta2,
    meta3,
    CLUSTER
"""


def fetch_store_cut_size_from_snowflake():
    conn = _snowflake_connect()
    try:
        cur = conn.cursor()
        print("Fetching Store Cut Size from offline_master_daily_report_1...")
        cur.execute(STORE_CUT_SIZE_QUERY)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    def _f(v):
        return float(v) if v is not None else None

    store_cut_size = []
    for row in rows:
        (l1, cat, m1, m2, m3, cluster, total_store_count_in_cluster,
         unique_option_count, store_count_with_cut_size, pct_cut, pct_full) = row
        store_cut_size.append({
            "l1": l1, "cat": cat, "m1": m1, "m2": m2, "m3": m3, "cluster": cluster,
            "total_store_count_in_cluster": int(total_store_count_in_cluster or 0),
            "unique_option_count": int(unique_option_count or 0),
            "store_count_with_cut_size": int(store_count_with_cut_size or 0),
            "pct_cut": _f(pct_cut), "pct_full": _f(pct_full),
        })
    print(f"  -> {len(store_cut_size)} store_cut_size rows from Snowflake")
    return store_cut_size


# ---------------------------------------------------------------------------
# Snowflake: Sales vs Inwards size curve (Sales vs Inwards tab). Per L1 x
# Category x Meta1-3 x Month x Size, absolute Inward qty and Sales qty by
# channel (Marketplace/Shopify/Store), over the trailing ~7 months (6 full
# months plus the current in-progress one -- the tab drops the partial
# current month client-side, same treatment as the Clicks trend chart).
# ---------------------------------------------------------------------------
SALES_VS_INWARDS_QUERY = r"""
WITH

meta_map_sales AS (
    SELECT
        UPPER(
            IFF(
                UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE 'MP%'
                OR UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE '4C-%',
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '-.*$', '')
            )
        ) AS sku_group_clean,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.L1_CATEGORY)) = 'LONG TAIL' THEN a.L1_CATEGORY END),
            MIN(CASE WHEN UPPER(TRIM(a.L1_CATEGORY)) = 'PLUS' THEN a.L1_CATEGORY END),
            MIN(CASE WHEN UPPER(TRIM(a.L1_CATEGORY)) = 'LUXE' THEN a.L1_CATEGORY END),
            MIN(CASE WHEN UPPER(TRIM(a.L1_CATEGORY)) = 'SNITCH' THEN a.L1_CATEGORY END),
            MIN(a.L1_CATEGORY)
        ) AS l1_category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'SHIRTS' THEN a.CATEGORY END),
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'TSHIRTS' THEN a.CATEGORY END),
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'JEANS' THEN a.CATEGORY END),
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'TROUSERS' THEN a.CATEGORY END),
            MIN(a.CATEGORY)
        ) AS category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'SHIRTS' THEN a.META1 END),
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'TSHIRTS' THEN a.META1 END),
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'JEANS' THEN a.META1 END),
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'TROUSERS' THEN a.META1 END),
            MIN(a.META1)
        ) AS meta1,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'SHIRTS' THEN a.META2 END),
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'TSHIRTS' THEN a.META2 END),
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'JEANS' THEN a.META2 END),
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'TROUSERS' THEN a.META2 END),
            MIN(a.META2)
        ) AS meta2,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'SHIRTS' THEN a.META3 END),
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'TSHIRTS' THEN a.META3 END),
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'JEANS' THEN a.META3 END),
            MIN(CASE WHEN UPPER(TRIM(a.CATEGORY)) = 'TROUSERS' THEN a.META3 END),
            MIN(a.META3)
        ) AS meta3
    FROM snitch_db.maplemonk.meta_mapping_cogs_sku_2 a
    GROUP BY 1
),

sales_base AS (
    SELECT
        h.date,
        UPPER(
            IFF(
                UPPER(REPLACE(h.SKU_GROUP,' ','')) LIKE 'MP%'
                OR UPPER(REPLACE(h.SKU_GROUP,' ','')) LIKE '4C-%',
                REGEXP_REPLACE(REPLACE(h.SKU_GROUP,' ',''), '^([^-]+-[^-]+).*$', '\1'),
                REGEXP_REPLACE(REPLACE(h.SKU_GROUP,' ',''), '-.*$', '')
            )
        ) AS sku_group_clean,
        CASE UPPER(TRIM(h."SIZE"))
            WHEN '2XL' THEN 'XXL'
            WHEN 'XXXL' THEN '3XL'
            ELSE UPPER(TRIM(h."SIZE"))
        END AS size_clean,
        h.TYPE AS channel_type,
        SUM(h.gross_quantity) AS qty
    FROM snitch_db.maplemonk.horizontal_sales_categories h
    WHERE h.gross_quantity > 0
      AND UPPER(h.SKU_GROUP) NOT LIKE '%CB%'
      AND h.date >= DATEADD('month', -6, DATE_TRUNC('month', CURRENT_DATE()))
      AND h.date < DATE_TRUNC('month', CURRENT_DATE()) + INTERVAL '1 month'
      AND TRIM(COALESCE(h."SIZE", '')) <> ''
    GROUP BY 1,2,3,4
),

sales_joined AS (
    SELECT
        mm.l1_category,
        mm.category,
        mm.meta1,
        mm.meta2,
        CASE
            WHEN mm.category = 'shirts' THEN
                CASE LOWER(TRIM(mm.meta3))
                    WHEN 'custom fit'   THEN 'slim fit'
                    WHEN 'loose fit'    THEN 'box fit'
                    WHEN 'straight fit' THEN 'regular fit'
                    ELSE mm.meta3
                END
            ELSE mm.meta3
        END AS meta3,
        sb.size_clean AS size,
        sb.channel_type,
        sb.qty,
        TO_CHAR(sb.date, 'Mon YYYY') AS month_label,
        DATE_TRUNC('month', sb.date) AS month_sort
    FROM sales_base sb
    JOIN meta_map_sales mm ON mm.sku_group_clean = sb.sku_group_clean
    WHERE UPPER(TRIM(REPLACE(mm.l1_category, '_', ' '))) <> 'LONG TAIL'
      AND NOT (
            LOWER(TRIM(mm.category)) IN ('shirts','tshirts','jeans','trousers')
            AND (mm.meta1 IS NULL OR TRIM(mm.meta1) = '')
          )
),

sales_grain_agg AS (
    SELECT
        l1_category, category, meta1, meta2, meta3,
        month_label, month_sort, size, channel_type,
        SUM(qty) AS qty
    FROM sales_joined
    GROUP BY 1,2,3,4,5,6,7,8,9
),

sales_pivot AS (
    SELECT
        l1_category, category, meta1, meta2, meta3,
        month_label, month_sort, size,
        SUM(CASE WHEN channel_type = 'Marketplace' THEN qty ELSE 0 END) AS "Marketplace_Qty",
        SUM(CASE WHEN channel_type = 'Shopify' THEN qty ELSE 0 END) AS "Shopify_Qty",
        SUM(CASE WHEN channel_type = 'Store' THEN qty ELSE 0 END) AS "Store_Qty"
    FROM sales_grain_agg
    GROUP BY 1,2,3,4,5,6,7,8
),

putaway_raw AS (
    SELECT
        UPPER(
            TRIM(
                REVERSE(
                    SUBSTRING(
                        REVERSE("Item Type skuCode"),
                        CHARINDEX('-', REVERSE("Item Type skuCode")) + 1,
                        LEN("Item Type skuCode")
                    )
                )
            )
        ) AS SKU_group,
        CASE TRIM(
                RIGHT(
                    "Item Type skuCode",
                    CHARINDEX('-', REVERSE("Item Type skuCode")) - 1
                )
             )
            WHEN '28' THEN 'xs'
            WHEN '30' THEN 's'
            WHEN '32' THEN 'm'
            WHEN '34' THEN 'l'
            WHEN '36' THEN 'xl'
            WHEN '38' THEN 'xxl'
            WHEN '40' THEN '3xl'
            WHEN '42' THEN '4xl'
            WHEN '44' THEN '5xl'
            WHEN '46' THEN '6xl'
            WHEN '48' THEN '7xl'
            WHEN '50' THEN '8xl'
            ELSE TRIM(
                    RIGHT(
                        "Item Type skuCode",
                        CHARINDEX('-', REVERSE("Item Type skuCode")) - 1
                    )
                 )
        END AS size,
        LOWER(FINAL_TYPE) AS type,
        "PUTAWAY_UPDATED"::DATE AS date,
        SUM("PUTAWAY_COMPLETED_QUANTITY") AS Qty
    FROM snitch_db.maplemonk.putaway_tracking
    WHERE "PUTAWAY_UPDATED"::DATE >= DATEADD('month', -6, DATE_TRUNC('month', CURRENT_DATE()))
      AND "PUTAWAY_UPDATED"::DATE < DATE_TRUNC('month', CURRENT_DATE()) + INTERVAL '1 month'
      AND LOWER(FINAL_TYPE) LIKE 'new%'
    GROUP BY 1,2,3,4
),

meta_map_inward AS (
    SELECT
        UPPER(
            IFF(
                UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE 'MP%'
                OR UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE '4C-%',
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '-.*$', '')
            )
        ) AS sku_group_clean,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LONG TAIL' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'PLUS' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LUXE' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'SNITCH' THEN a.l1_category END),
            MIN(a.l1_category)
        ) AS l1_category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.category END),
            MIN(a.category)
        ) AS category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta1 END),
            MIN(a.meta1)
        ) AS meta1,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta2 END),
            MIN(a.meta2)
        ) AS meta2,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta3 END),
            MIN(a.meta3)
        ) AS meta3
    FROM snitch_db.maplemonk.meta_mapping_cogs_sku_2 a
    GROUP BY 1
),

inward_joined AS (
    SELECT
        mm.l1_category,
        mm.category,
        mm.meta1,
        mm.meta2,
        CASE
            WHEN mm.category = 'shirts' THEN
                CASE LOWER(TRIM(mm.meta3))
                    WHEN 'custom fit'   THEN 'slim fit'
                    WHEN 'loose fit'    THEN 'box fit'
                    WHEN 'straight fit' THEN 'regular fit'
                    ELSE mm.meta3
                END
            ELSE mm.meta3
        END AS meta3,
        p.size,
        p.qty,
        TO_CHAR(p.date, 'Mon YYYY') AS month_label,
        DATE_TRUNC('month', p.date) AS month_sort
    FROM putaway_raw p
    INNER JOIN meta_map_inward mm
        ON mm.sku_group_clean = UPPER(
            IFF(
                UPPER(REPLACE(p.SKU_GROUP, ' ', '')) LIKE 'MP%'
                OR UPPER(REPLACE(p.SKU_GROUP, ' ', '')) LIKE '4C-%',
                REGEXP_REPLACE(REPLACE(p.SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
                REGEXP_REPLACE(REPLACE(p.SKU_GROUP, ' ', ''), '-.*$', '')
            )
        )
    WHERE UPPER(TRIM(REPLACE(mm.l1_category, '_', ' '))) <> 'LONG TAIL'
      AND NOT (
            LOWER(TRIM(mm.category)) IN ('shirts','tshirts','jeans','trousers')
            AND (mm.meta1 IS NULL OR TRIM(mm.meta1) = '')
          )
),

inward_grain_agg AS (
    SELECT
        l1_category,
        category,
        meta1,
        meta2,
        meta3,
        month_label,
        month_sort,
        CASE
            WHEN size IS NULL OR TRIM(size) = '' THEN 'L'
            WHEN size = '01M' THEN 'M'
            WHEN size = '01XL' THEN 'XL'
            WHEN size = '06XL' THEN '6XL'
            WHEN UPPER(size) IN ('XS', 'S', 'M', 'L', 'XL') THEN UPPER(size)
            WHEN UPPER(size) IN ('XXL', '2XL', 'XXU') THEN 'XXL'
            WHEN UPPER(size) IN ('XXXL', '3XL') THEN '3XL'
            WHEN UPPER(size) IN ('4XL', '5XL', '6XL', '7XL', '8XL') THEN UPPER(size)
            WHEN size = '28' THEN 'XS'
            WHEN size = '30' THEN 'S'
            WHEN size = '32' THEN 'M'
            WHEN size = '34' THEN 'L'
            WHEN size = '36' THEN 'XL'
            WHEN size = '38' THEN 'XXL'
            WHEN size = '40' THEN '3XL'
            WHEN size = '42' THEN '4XL'
            WHEN size = '44' THEN '5XL'
            WHEN size = '46' THEN '6XL'
            WHEN size = '48' THEN '7XL'
            WHEN size = '50' THEN '8XL'
            WHEN size = '39' THEN 'XXL'
            WHEN size = '41' THEN '4XL'
            WHEN size = '43' THEN '5XL'
            WHEN size = '45' THEN '6XL'
            WHEN size = '1' OR size = '01' THEN 'XS'
            WHEN size = '2' OR size = '02' THEN 'S'
            WHEN size = '3' OR size = '03' THEN 'M'
            WHEN size = '4' OR size = '04' THEN 'L'
            WHEN size = '5' OR size = '05' THEN 'XL'
            WHEN size = '6' OR size = '06' THEN 'XXL'
            ELSE UPPER(size)
        END AS size,
        qty
    FROM inward_joined
),

inward_pivot AS (
    SELECT
        l1_category, category, meta1, meta2, meta3,
        month_label, month_sort, size,
        SUM(qty) AS "Inward_Qty"
    FROM inward_grain_agg
    GROUP BY 1,2,3,4,5,6,7,8
),

combined AS (
    SELECT
        COALESCE(s.l1_category, i.l1_category)   AS l1_category,
        COALESCE(s.category, i.category)         AS category,
        COALESCE(s.meta1, i.meta1)               AS meta1,
        COALESCE(s.meta2, i.meta2)               AS meta2,
        COALESCE(s.meta3, i.meta3)               AS meta3,
        COALESCE(s.month_label, i.month_label)   AS month,
        COALESCE(s.month_sort, i.month_sort)     AS month_sort,
        COALESCE(s.size, i.size)                 AS size,
        s."Marketplace_Qty",
        s."Shopify_Qty",
        s."Store_Qty",
        i."Inward_Qty"
    FROM sales_pivot s
    FULL OUTER JOIN inward_pivot i
        ON  s.l1_category IS NOT DISTINCT FROM i.l1_category
        AND s.category    IS NOT DISTINCT FROM i.category
        AND s.meta1       IS NOT DISTINCT FROM i.meta1
        AND s.meta2       IS NOT DISTINCT FROM i.meta2
        AND s.meta3       IS NOT DISTINCT FROM i.meta3
        AND s.month_sort  IS NOT DISTINCT FROM i.month_sort
        AND s.size        IS NOT DISTINCT FROM i.size
)

SELECT
    l1_category,
    category,
    meta1,
    meta2,
    meta3,
    month,
    month_sort,
    size,
    CASE
        WHEN LOWER(TRIM(l1_category)) IN ('snitch','luxe') AND size IN ('3XL','4XL','5XL','6XL','7XL','8XL') THEN 0
        WHEN LOWER(TRIM(l1_category)) = 'plus' AND size IN ('XS','S','M','L','XL','XXL') THEN 0
        ELSE "Marketplace_Qty"
    END AS "Marketplace_Qty",
    CASE
        WHEN LOWER(TRIM(l1_category)) IN ('snitch','luxe') AND size IN ('3XL','4XL','5XL','6XL','7XL','8XL') THEN 0
        WHEN LOWER(TRIM(l1_category)) = 'plus' AND size IN ('XS','S','M','L','XL','XXL') THEN 0
        ELSE "Shopify_Qty"
    END AS "Shopify_Qty",
    CASE
        WHEN LOWER(TRIM(l1_category)) IN ('snitch','luxe') AND size IN ('3XL','4XL','5XL','6XL','7XL','8XL') THEN 0
        WHEN LOWER(TRIM(l1_category)) = 'plus' AND size IN ('XS','S','M','L','XL','XXL') THEN 0
        ELSE "Store_Qty"
    END AS "Store_Qty",
    CASE
        WHEN LOWER(TRIM(l1_category)) IN ('snitch','luxe') AND size IN ('3XL','4XL','5XL','6XL','7XL','8XL') THEN 0
        WHEN LOWER(TRIM(l1_category)) = 'plus' AND size IN ('XS','S','M','L','XL','XXL') THEN 0
        ELSE "Inward_Qty"
    END AS "Inward_Qty"
FROM combined
ORDER BY l1_category, category, meta1, meta2, meta3, month_sort, size
"""


def fetch_sales_vs_inwards_from_snowflake():
    conn = _snowflake_connect()
    try:
        cur = conn.cursor()
        print("Fetching Sales vs Inwards size curve from horizontal_sales_categories + putaway_tracking...")
        cur.execute(SALES_VS_INWARDS_QUERY)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    sales_vs_inwards = []
    for row in rows:
        (l1, cat, m1, m2, m3, month, month_sort, size,
         mp_qty, shopify_qty, store_qty, inward_qty) = row
        month_key = month_sort.strftime("%Y-%m") if hasattr(month_sort, "strftime") else None
        sales_vs_inwards.append({
            "l1": l1, "cat": cat, "m1": m1, "m2": m2, "m3": m3,
            "month": month, "month_key": month_key, "size": size,
            "mp_qty": float(mp_qty or 0), "shopify_qty": float(shopify_qty or 0),
            "store_qty": float(store_qty or 0), "inward_qty": float(inward_qty or 0),
        })
    print(f"  -> {len(sales_vs_inwards)} sales_vs_inwards rows from Snowflake")
    return sales_vs_inwards


# ---------------------------------------------------------------------------
# Snowflake: Store Returns (Store Returns tab). Store-return putaways
# (FINAL_TYPE like 'store return%', distinct from 'new%' inwards elsewhere in
# this file) at SKU-group grain, monthly, over the trailing 6 full months --
# the query's own WHERE clause already excludes the current in-progress month.
# ---------------------------------------------------------------------------
STORE_RETURNS_QUERY = r"""
with a as (
    SELECT
        upper(
            trim(
                REVERSE(
                    SUBSTRING(
                        REVERSE("Item Type skuCode"),
                        CHARINDEX('-', REVERSE("Item Type skuCode")) + 1,
                        LEN("Item Type skuCode")
                    )
                )
            )
        ) AS SKU_group,
        LOWER(FINAL_TYPE) type,
        DATE_TRUNC('MONTH', "PUTAWAY_UPDATED"::DATE) as month,
        SUM("PUTAWAY_COMPLETED_QUANTITY") AS Qty
    FROM
        snitch_db.maplemonk.putaway_tracking
    WHERE
        "PUTAWAY_UPDATED"::DATE >= DATEADD(MONTH, -6, DATE_TRUNC('MONTH', CURRENT_DATE))
        AND "PUTAWAY_UPDATED"::DATE < DATE_TRUNC('MONTH', CURRENT_DATE)
        AND LOWER(FINAL_TYPE) like 'store return%'
    GROUP BY
        1,2,3
),
meta_map AS (
    SELECT
       UPPER(
            IFF(
                UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE 'MP%'
                OR UPPER(REPLACE(SKU_GROUP, ' ', '')) LIKE '4C-%',
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
                REGEXP_REPLACE(REPLACE(SKU_GROUP, ' ', ''), '-.*$', '')
            )
        ) AS sku_group_clean,
       COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LONG TAIL' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'PLUS' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LUXE' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'SNITCH' THEN a.l1_category END),
            MIN(a.l1_category)
        ) AS l1_category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.category END),
            MIN(a.category)
        ) AS category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta1 END),
            MIN(a.meta1)
        ) AS meta1,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta2 END),
            MIN(a.meta2)
        ) AS meta2,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta3 END),
            MIN(a.meta3)
        ) AS meta3,
        max(cogs) as cogs
    FROM snitch_db.maplemonk.meta_mapping_cogs_sku a
    GROUP BY 1
),
lkp_agg AS (
    SELECT
        TRIM(UPPER(SKU_GROUP)) AS sku_group_lkp,
        MAX(FINAL_LIVE_DATE) AS FINAL_LIVE_DATE
    FROM snitch_db.maplemonk.base_product
    GROUP BY 1
)
select
    a.month,
    a.sku_group,
    lkp_agg.FINAL_LIVE_DATE,
    meta_map.l1_category,
    meta_map.category,
    meta_map.meta1,
    meta_map.meta2,
    meta_map.meta3,
    sum(a.qty*meta_map.cogs) cogs_value,
    sum(a.qty) qty
from a
left join meta_map on meta_map.sku_group_clean = UPPER(
        IFF(
            UPPER(REPLACE(a.SKU_GROUP, ' ', '')) LIKE 'MP%'
            OR UPPER(REPLACE(a.SKU_GROUP, ' ', '')) LIKE '4C-%',
            REGEXP_REPLACE(REPLACE(a.SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
            REGEXP_REPLACE(REPLACE(a.SKU_GROUP, ' ', ''), '-.*$', '')
        )
    )
left join lkp_agg on lkp_agg.sku_group_lkp = TRIM(UPPER(a.SKU_GROUP))
group by 1,2,3,4,5,6,7,8
order by 1
"""


def fetch_store_returns_from_snowflake():
    conn = _snowflake_connect()
    try:
        cur = conn.cursor()
        print("Fetching Store Returns from putaway_tracking...")
        cur.execute(STORE_RETURNS_QUERY)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    def _d(v):
        if v is None:
            return None
        return v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v)[:10]

    store_returns = []
    for row in rows:
        (month, sku_group, final_live_date, l1, cat, m1, m2, m3, cogs_value, qty) = row
        month_key = month.strftime("%Y-%m") if hasattr(month, "strftime") else None
        store_returns.append({
            "month_key": month_key, "sku_group": sku_group, "live_date": _d(final_live_date),
            "l1": l1, "cat": cat, "m1": m1, "m2": m2, "m3": m3,
            "cogs_value": float(cogs_value or 0), "qty": float(qty or 0),
        })
    print(f"  -> {len(store_returns)} store_returns rows from Snowflake")
    return store_returns


# ---------------------------------------------------------------------------
# Snowflake: daily-grain actual Returns (RETURNS_DATA), joined to the product
# master (meta_mapping_cogs_sku) to get L1/Category/Meta1/2/3 per SKU_GROUP,
# for the new Analytics tab's "Returns" metric (actual, not the learned/frozen
# % used by the existing Returns tab).
#
# Channel split: RETURNS_DATA only has OVERALL_RETURNS_QTY and
# SHOPIFY_RETURNS_QTY (no separate Marketplace/Offline columns). Per Aditya:
# Offline returns are always 0, so:
#   shopify_returns_qty    = SHOPIFY_RETURNS_QTY (exact)
#   offline_returns_qty    = 0 (fixed)
#   marketplace_returns_qty = OVERALL_RETURNS_QTY - SHOPIFY_RETURNS_QTY (exact,
#                             since Offline contributes nothing to the gap)
# One row per (date, sku_group) -- exploded into per-channel qty here, at the
# SAME leaf grain (l1/cat/m1/m2/m3) as Sales, so it can be sliced identically.
# ---------------------------------------------------------------------------
RETURNS_QUERY = rf"""
WITH meta_map AS (
    SELECT
        UPPER(
            IFF(
                UPPER(REPLACE(a.SKU_GROUP, ' ', '')) LIKE 'MP%'
                OR UPPER(REPLACE(a.SKU_GROUP, ' ', '')) LIKE '4C-%',
                REGEXP_REPLACE(REPLACE(a.SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
                REGEXP_REPLACE(REPLACE(a.SKU_GROUP, ' ', ''), '-.*$', '')
            )
        ) AS sku_group_clean,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LONG TAIL' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'PLUS' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'LUXE' THEN a.l1_category END),
            MIN(CASE WHEN UPPER(TRIM(a.l1_category)) = 'SNITCH' THEN a.l1_category END),
            MIN(a.l1_category)
        ) AS l1_category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.category END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.category END),
            MIN(a.category)
        ) AS category,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta1 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta1 END),
            MIN(a.meta1)
        ) AS meta1,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta2 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta2 END),
            MIN(a.meta2)
        ) AS meta2,
        COALESCE(
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'SHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TSHIRTS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'JEANS' THEN a.meta3 END),
            MIN(CASE WHEN UPPER(TRIM(a.category)) = 'TROUSERS' THEN a.meta3 END),
            MIN(a.meta3)
        ) AS meta3,
        MAX(a.cogs) AS cogs
    FROM snitch_db.maplemonk.meta_mapping_cogs_sku a
    GROUP BY 1
),
returns_with_parent AS (
    SELECT
        r.*,
        UPPER(
            IFF(
                UPPER(REPLACE(r.SKU_GROUP, ' ', '')) LIKE 'MP%'
                OR UPPER(REPLACE(r.SKU_GROUP, ' ', '')) LIKE '4C-%',
                REGEXP_REPLACE(REPLACE(r.SKU_GROUP, ' ', ''), '^([^-]+-[^-]+).*$', '\1'),
                REGEXP_REPLACE(REPLACE(r.SKU_GROUP, ' ', ''), '-.*$', '')
            )
        ) AS sku_group_clean
    FROM SNITCH_DB.MAPLEMONK.RETURNS_DATA r
    WHERE r.DATE >= '{ROLLING_CUTOFF_DATE}'
)
SELECT
    r.DATE,
    m.l1_category, m.category, m.meta1, m.meta2, m.meta3,
    SUM(r.SHOPIFY_REVENUE) AS SHOPIFY_REVENUE, SUM(r.SHOPIFY_QTY) AS SHOPIFY_QTY,
    SUM(r.MP_REVENUE) AS MP_REVENUE, SUM(r.MP_QTY) AS MP_QTY,
    SUM(r.OFFLINE_REVENUE) AS OFFLINE_REVENUE, SUM(r.OFFLINE_QTY) AS OFFLINE_QTY,
    SUM(r.OVERALL_RETURNS_QTY) AS OVERALL_RETURNS_QTY, SUM(r.SHOPIFY_RETURNS_QTY) AS SHOPIFY_RETURNS_QTY
FROM returns_with_parent r
LEFT JOIN meta_map m
    ON r.sku_group_clean = m.sku_group_clean
GROUP BY r.DATE, m.l1_category, m.category, m.meta1, m.meta2, m.meta3
"""


def fetch_returns_from_snowflake():
    conn = _snowflake_connect()
    try:
        cur = conn.cursor()
        print("Fetching returns from RETURNS_DATA (joined to meta_mapping_cogs_sku)...")
        cur.execute(RETURNS_QUERY)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    returns_actual = []
    unmapped = 0
    for row in rows:
        (dt, l1, cat, m1, m2, m3,
         shopify_rev, shopify_qty, mp_rev, mp_qty, offline_rev, offline_qty,
         overall_returns_qty, shopify_returns_qty) = row
        if l1 is None or cat is None:
            unmapped += 1
            continue
        d = dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt)[:10]
        overall_ret = float(overall_returns_qty or 0)
        shopify_ret = float(shopify_returns_qty or 0)
        mp_ret = max(0.0, overall_ret - shopify_ret)  # offline always 0, per Aditya
        base = {"date": d, "month": month_key(dt), "l1": l1, "cat": cat, "m1": m1, "m2": m2, "m3": m3}
        if float(shopify_qty or 0) or shopify_ret:
            returns_actual.append({**base, "channel": "Shopify",
                                    "sales_qty": float(shopify_qty or 0), "sales_value": float(shopify_rev or 0),
                                    "return_qty": shopify_ret})
        if float(mp_qty or 0) or mp_ret:
            returns_actual.append({**base, "channel": "Marketplace",
                                    "sales_qty": float(mp_qty or 0), "sales_value": float(mp_rev or 0),
                                    "return_qty": mp_ret})
        if float(offline_qty or 0):
            returns_actual.append({**base, "channel": "Offline",
                                    "sales_qty": float(offline_qty or 0), "sales_value": float(offline_rev or 0),
                                    "return_qty": 0.0})
    if unmapped:
        print(f"  NOTE: {unmapped} rows had no L1/Category match in meta_mapping_cogs_sku, skipped.")
    print(f"  -> {len(returns_actual)} returns rows (per-channel) from Snowflake")
    return returns_actual


def month_key(dt):
    if dt is None:
        return None
    if isinstance(dt, str):
        return None  # e.g. "No Date mentioned" -- treat as undated, not a real month
    if hasattr(dt, "year") and dt.year < 2020:
        return None  # placeholder/epoch-error dates (e.g. 1970-01-01) -- not real
    return dt.strftime("%Y-%m")


def build_returns():
    """Loads the frozen, one-time 'learning' dataset (data/returns_learned.json,
    monthly Sales Qty / Return Qty at Channel x L1 x Category grain, 2026 only,
    MP-SOR already merged into Marketplace). This is NOT re-derived from a live
    source -- it's a static snapshot used as the starting point for the editable
    Returns tab in the dashboard, which is where the actually-applied rates
    (with caps, and any manual overrides) live from here on."""
    path = os.path.join(HERE, "data", "returns_learned.json")
    if not os.path.exists(path):
        print(f"NOTE: {path} not found -- Returns tab will start empty (0% assumed everywhere).")
        return []
    with open(path) as f:
        learned = json.load(f)
    print(f"Returns: loaded {len(learned)} frozen monthly learning rows.")
    return learned


def build_channel_mix_pct():
    """Computes OLD channel-mix % per (L1, Category, Month) from the 4
    Planned_Qty_only files -- summed across Meta/ASP-Bin/store, kept distinct
    per Channel (Warehouse excluded, not a sales channel)."""
    old_qty = {}  # (l1, cat, channel, month) -> qty
    any_found = False
    for fname in TARGET_FILES:
        path = os.path.join(TARGETS_DIR, fname)
        if not os.path.exists(path):
            print(f"NOTE: target file not found (skipping): data/targets/{fname}")
            continue
        any_found = True
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb["Sheet1"]
        col_idx = None
        for row in ws.iter_rows(values_only=True):
            if col_idx is None:
                if row and "L1_CATEGORY" in row:
                    col_idx = {name: i for i, name in enumerate(row)}
                continue
            l1 = row[col_idx["L1_CATEGORY"]]
            if l1 is None or l1 == "L1_CATEGORY":
                continue
            cat = (row[col_idx["CATEGORY"]] or "").strip().lower()
            channel_raw = row[col_idx["type"]]
            if channel_raw is None or channel_raw == "type" or channel_raw == "Warehouse":
                continue
            channel = TARGET_CHANNEL_MAP.get(channel_raw, channel_raw)
            for col_name, month_iso in MONTH_COL_MAP.items():
                idx = col_idx.get(col_name)
                if idx is None:
                    continue
                val = row[idx] or 0
                key = (l1, cat, channel, month_iso)
                old_qty[key] = old_qty.get(key, 0) + val
        wb.close()
    if not any_found:
        print("NOTE: no target files found in data/targets/ -- cannot compute channel mix.")
        return {}

    # totals per (l1, cat, month) across all channels
    totals = {}
    for (l1, cat, channel, month), qty in old_qty.items():
        tkey = (l1, cat, month)
        totals[tkey] = totals.get(tkey, 0) + qty

    pct = {}
    for (l1, cat, channel, month), qty in old_qty.items():
        total = totals.get((l1, cat, month), 0)
        if total > 0:
            pct[(l1, cat, channel, month)] = qty / total
    return pct


def build_new_totals():
    """Reads data/targets_new_total.xlsx (the new all-channel-combined Planned
    Qty numbers), aggregating away Meta1/2/3 and ASP Bin (ASP is parked for now)
    down to L1 x Category x Month."""
    if not os.path.exists(NEW_TOTAL_FILE):
        print("NOTE: data/targets_new_total.xlsx not found -- no new target totals to apply.")
        return {}
    wb = openpyxl.load_workbook(NEW_TOTAL_FILE, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    col_idx = None
    NEW_MONTH_COLS = {
        "Planned Qty AUG'26": "2026-08", "Planned Qty SEP'26": "2026-09",
        "Planned Qty OCT'26": "2026-10", "Planned Qty NOV'26": "2026-11",
        "Planned Qty DEC'26": "2026-12",
    }
    totals = {}
    for row in ws.iter_rows(values_only=True):
        if col_idx is None:
            if row and "Category" in row and "L1" in row:
                col_idx = {name: i for i, name in enumerate(row)}
            continue
        cat_raw = row[col_idx["Category"]]
        if cat_raw is None:
            continue
        cat = str(cat_raw).strip().lower()
        l1 = row[col_idx["L1"]]
        for col_name, month_iso in NEW_MONTH_COLS.items():
            idx = col_idx.get(col_name)
            if idx is None:
                continue
            val = row[idx] or 0
            key = (l1, cat, month_iso)
            totals[key] = totals.get(key, 0) + val
    return totals


def build_targets():
    """New target logic: preserves the OLD channel-mix % (from the 4
    Planned_Qty_only files) and applies it to the NEW total (from
    data/targets_new_total.xlsx), per L1 x Category x Month. The old files
    are now used only to derive the split ratio, not as target values
    themselves -- the new file's totals are authoritative."""
    pct = build_channel_mix_pct()
    new_totals = build_new_totals()

    targets = []
    for (l1, cat, month), new_total in new_totals.items():
        matching_channels = [k for k in pct if k[0] == l1 and k[1] == cat and k[3] == month]
        if not matching_channels:
            continue  # no old channel-mix data for this L1+Cat+Month -- can't split honestly
        for key in matching_channels:
            _, _, channel, _ = key
            share = pct[key]
            targets.append({"channel": channel, "l1": l1, "cat": cat, "month": month, "qty": new_total * share})
    return targets


def build(xlsx_path_or_workbook):
    if isinstance(xlsx_path_or_workbook, str):
        wb = openpyxl.load_workbook(xlsx_path_or_workbook, data_only=True)
    else:
        wb = xlsx_path_or_workbook  # already a loaded workbook (or SimpleWorkbook from the API path)

    # ---- Sales (now from Snowflake, daily grain -- see fetch_sales_from_snowflake) ----
    sales = fetch_sales_from_snowflake()
    sales_by_sku = fetch_sales_by_sku_from_snowflake()
    returns_by_sku = fetch_returns_by_sku_from_snowflake()
    returns_actual = fetch_returns_from_snowflake()
    images = fetch_images_from_snowflake()
    live_sku_groups = fetch_live_sku_groups_from_snowflake()
    clicks = fetch_clicks_from_snowflake()
    clicks_monthly = fetch_clicks_monthly_from_snowflake()
    store_cut_size = fetch_store_cut_size_from_snowflake()
    sales_vs_inwards = fetch_sales_vs_inwards_from_snowflake()
    store_returns = fetch_store_returns_from_snowflake()

    # ---- Inventory ----
    # Kept at DAILY grain (not summed to month) because Closing(month) = the
    # latest dated snapshot within that month, per leaf combo -- summing days
    # together would wildly overstate stock. The frontend does the "latest
    # snapshot in month" pick, exactly matching the Overall tab's
    # SUMIFS(...,'CI v2'!date, MAXIFS(...)) formula.
    ws = wb["Inv Data 2"]
    inventory = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        inv_date, l1, cat, m1, m2, m3, mp_qty, online_qty, offline_qty, total_qty, mp_val, online_val, offline_val, total_val = row[:14]
        if l1 is None or cat is None:
            # stray/blank rows found in the source sheet (no L1 or Category) -- not real
            # inventory, skip so they don't pollute L1/Category rollups.
            continue
        d = inv_date.strftime("%Y-%m-%d") if hasattr(inv_date, "strftime") else str(inv_date)[:10]
        inventory.append({
            "date": d, "l1": l1, "cat": cat, "m1": m1, "m2": m2, "m3": m3,
            "mp_qty": mp_qty or 0, "online_qty": online_qty or 0, "offline_qty": offline_qty or 0,
            "total_qty": total_qty or 0,
            "mp_value": mp_val or 0, "online_value": online_val or 0, "offline_value": offline_val or 0,
            "total_value": total_val or 0
        })

    # ---- Pipeline (for Inwards) ----
    # Matched at the same leaf grain as inventory (L1+Cat+Meta1+Meta2+Meta3), keyed
    # by FDD_MONTH -- the month a batch is expected to land. No STATUS filter,
    # matching the original Overall-tab formula exactly.
    ws = wb["Pipeline"]
    pipeline = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        l1, cat, m1, m2, m3 = row[9], row[10], row[11], row[12], row[13]
        fdd_month, qty, cogs_unit = row[21], row[22], row[23]
        if l1 is None or cat is None:
            continue
        qty = qty or 0
        pipeline.append({
            "l1": l1, "cat": cat, "m1": m1, "m2": m2, "m3": m3,
            "fdd_month": month_key(fdd_month),  # None if undated
            "qty": qty,
            "cogs_value": qty * (cogs_unit or 0)
        })

    targets = build_targets()
    returns = build_returns()

    og_path = os.path.join(HERE, "data", "targets_og_fallback.json")
    targets_og = []
    if os.path.exists(og_path):
        with open(og_path) as f:
            targets_og = json.load(f)
    else:
        print("NOTE: data/targets_og_fallback.json not found -- no fallback for uncovered categories.")

    data_dir = os.path.dirname(OUT)
    os.makedirs(data_dir, exist_ok=True)

    # Split across multiple files instead of one giant data.json:
    #   1. GitHub hard-blocks any single file over 100MB -- one big file was
    #      already past that (225MB+) and would fail to push outright.
    #   2. The browser can fetch these in parallel and start rendering the
    #      Dashboard as soon as sales/inventory/meta arrive, without waiting
    #      on the (much larger) returns_actual/sales_by_sku to finish parsing.
    parts = {
        "meta.json": {
            "generated_at": datetime.datetime.now().isoformat(),
            "images": images,
        },
        "sales.json": {"sales": sales},
        "sales_by_sku.json": {"sales_by_sku": sales_by_sku, "returns_by_sku": returns_by_sku},
        "inventory.json": {"inventory": inventory},
        "pipeline.json": {"pipeline": pipeline},
        "returns_actual.json": {"returns_actual": returns_actual},
        "livesku.json": {"live_sku_groups": live_sku_groups},
        "clicks.json": {"clicks": clicks, "clicks_monthly": clicks_monthly},
        "storecutsize.json": {"store_cut_size": store_cut_size},
        "salesvsinwards.json": {"sales_vs_inwards": sales_vs_inwards},
        "storereturns.json": {"store_returns": store_returns},
        "targets.json": {
            "targets": targets,        # primary source: the 4 category files, multi-channel
            "targets_og": targets_og,  # fallback: old Targets tab, Shopify-only, all categories
            "returns": returns,        # frozen monthly learning rows: Channel x L1 x Cat x Month (2026)
        },
    }
    written = []
    for filename, payload in parts.items():
        path = os.path.join(data_dir, filename)
        with open(path, "w") as f:
            json.dump(payload, f)
        size_mb = os.path.getsize(path) / (1024*1024)
        written.append((filename, size_mb))

    print(f"sales rows: {len(sales)}, sales_by_sku rows: {len(sales_by_sku)}, returns_by_sku rows: {len(returns_by_sku)}, "
          f"inventory rows: {len(inventory)}, pipeline rows: {len(pipeline)}, "
          f"target rows (new files): {len(targets)}, target rows (OG fallback): {len(targets_og)}, "
          f"returns_actual rows: {len(returns_actual)}, images: {len(images)}, live_sku_groups rows: {len(live_sku_groups)}, "
          f"clicks rows: {len(clicks)}, clicks_monthly rows: {len(clicks_monthly)}, "
          f"store_cut_size rows: {len(store_cut_size)}, sales_vs_inwards rows: {len(sales_vs_inwards)}, "
          f"store_returns rows: {len(store_returns)}")
    if len(sales) > 0 and len(inventory) == 0:
        print("\n⚠️  WARNING: Sales has rows but Inventory is empty. This is the exact symptom")
        print("   of an unauthenticated download failing to resolve IMPORTRANGE on Inv Data 2.")
        print("   Fix: re-download Automation_Data.xlsx manually via your browser (File > Download")
        print("   > Microsoft Excel), then re-run with --local. See the module docstring for detail.\n")

    print("\nWrote data files:")
    total_mb = 0
    for filename, size_mb in written:
        flag = "  ⚠️ still >90MB, close to GitHub's 100MB limit!" if size_mb > 90 else ""
        print(f"  data/{filename}: {size_mb:.1f} MB{flag}")
        total_mb += size_mb
    print(f"  TOTAL: {total_mb:.1f} MB across {len(written)} files -> {data_dir}")


if __name__ == "__main__":
    if "--local" in sys.argv:
        if not os.path.exists(LOCAL_XLSX):
            print(f"ERROR: {LOCAL_XLSX} not found.")
            sys.exit(1)
        build(LOCAL_XLSX)
    elif "--export-url" in sys.argv:
        download_from_gsheet(SHEET_ID, LOCAL_XLSX)
        build(LOCAL_XLSX)
    else:
        wb = fetch_via_sheets_api(SHEET_ID)
        build(wb)